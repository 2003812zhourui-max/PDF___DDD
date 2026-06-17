import argparse
import base64
import csv
import getpass
import json
import os
import re
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlencode

try:
    from playwright.sync_api import Error as PlaywrightError
    from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright
except ModuleNotFoundError:
    PlaywrightError = Exception
    PlaywrightTimeoutError = TimeoutError
    Page = Any
    sync_playwright = None

try:
    import requests
except ModuleNotFoundError:
    requests = None

try:
    from pypdf import PdfReader
except ModuleNotFoundError:
    PdfReader = None


TARGET_PAGE = "https://omp.xlwms.com/wms/outbound/parcel"
API_ORIGIN = "https://omp.xlwms.com"
LIST_API = "/gateway/wms/blDelivery/page"
DETAIL_API = "/gateway/wms/blDelivery/detail"
DOWNLOAD_URL_API = "/gateway/wms/appendix/getPreviewAndDownLoadUrl"

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent
CURRENT_WORK_DIR = Path(os.getcwd()).resolve()
LOG_DIR = BASE_DIR / "logs"
PDF_DIR = BASE_DIR / "pdf_downloads"
LOG_FILE = LOG_DIR / "download_log.csv"
AUTH_STATE_FILE = BASE_DIR / "wms_storage_state.json"
PDF_VALIDATION_LOG_FILE = LOG_DIR / "pdf_validation_log.csv"
LABEL_METADATA_LOG_FILE = Path(os.environ.get("PDF_DDD_METADATA_JSONL") or BASE_DIR / "output" / "download_label_metadata.jsonl")
LABEL_METADATA_SUMMARY_FILE = Path(
    os.environ.get("PDF_DDD_METADATA_SUMMARY") or BASE_DIR / "output" / "download_label_metadata_summary.xlsx"
)
TRACK_SIGN_JS_CACHE: str | None = None

CSV_FIELDS = [
    "deliveryNo",
    "sourceNo",
    "expressNo",
    "customerCode",
    "whCode",
    "status",
    "filePath",
    "error",
    "downloadedAt",
]

PDF_VALIDATION_FIELDS = [
    "deliveryNo",
    "sourceNo",
    "expressNo",
    "customerCode",
    "whCode",
    "filePath",
    "pdfValid",
    "pageCount",
    "pageSizes",
    "detectedFormats",
    "formatOk",
    "error",
    "checkedAt",
]
LABEL_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}

FORBIDDEN_FETCH_HEADERS = {
    "accept-encoding",
    "connection",
    "content-length",
    "cookie",
    "host",
    "origin",
    "referer",
    "sec-fetch-dest",
    "sec-fetch-mode",
    "sec-fetch-site",
    "sec-fetch-user",
    "upgrade-insecure-requests",
    "user-agent",
}

MINIMAL_API_HEADER_NAMES = {
    "authorization",
    "x-requested-with",
}

TRACKING_COOKIE_HINTS = (
    "sensors",
    "analytics",
    "_ga",
    "_gid",
    "_gat",
    "hm_",
    "hmac",
    "ajs_",
    "amplitude",
    "track",
    "trace",
)


def log(message: str) -> None:
    print(f"[WMS批量下载] {message}", flush=True)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def resolve_storage_state_path(path_text: str) -> Path:
    candidates: list[Path] = []
    if path_text:
        candidates.append(Path(path_text).expanduser().resolve())
    candidates.append(AUTH_STATE_FILE.resolve())
    candidates.append((CURRENT_WORK_DIR / "wms_storage_state.json").resolve())

    unique_candidates: list[Path] = []
    seen: set[str] = set()
    for path in candidates:
        key = str(path).lower()
        if key not in seen:
            unique_candidates.append(path)
            seen.add(key)

    for path in unique_candidates:
        if path.exists():
            return path
    return unique_candidates[0]


def sanitize_filename_part(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text or "EMPTY"


def first_non_empty(record: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = record.get(key)
        if value is not None and value != "":
            return str(value)
    return ""


def reusable_headers(headers: dict[str, str]) -> dict[str, str]:
    result: dict[str, str] = {}
    for key, value in headers.items():
        lower_key = key.lower()
        if lower_key in FORBIDDEN_FETCH_HEADERS:
            continue
        if lower_key.startswith("sec-") or lower_key.startswith(":"):
            continue
        result[key] = value
    return result


def minimal_api_headers(headers: dict[str, str]) -> dict[str, str]:
    result = {"accept": "application/json, text/plain, */*"}
    for key, value in headers.items():
        lower_key = key.lower()
        if lower_key in MINIMAL_API_HEADER_NAMES:
            result[key] = value
    return result


def get_header(headers: dict[str, str], name: str) -> str:
    expected = name.lower()
    for key, value in headers.items():
        if key.lower() == expected:
            return value
    return ""


def limit_reached(count: int, limit: int) -> bool:
    return limit > 0 and count >= limit


def page_numbers(start: int, max_pages: int):
    page_no = start
    while max_pages <= 0 or page_no < start + max_pages:
        yield page_no
        page_no += 1


def is_tracking_cookie(cookie: dict[str, Any]) -> bool:
    name = str(cookie.get("name", "")).lower()
    value = str(cookie.get("value", ""))
    if any(hint in name for hint in TRACKING_COOKIE_HINTS):
        return True
    auth_name_hints = ("token", "auth", "session", "jwt", "login")
    return len(value) > 2000 and not any(hint in name for hint in auth_name_hints)


def cleanup_tracking_cookies(context) -> None:
    try:
        cookies = context.cookies()
    except Exception as exc:
        log(f"读取 Cookie 失败，跳过 Cookie 清理：{exc}")
        return

    removed = 0
    for cookie in cookies:
        domain = str(cookie.get("domain", ""))
        if "xlwms.com" not in domain:
            continue
        if not is_tracking_cookie(cookie):
            continue
        try:
            context.clear_cookies(
                name=cookie.get("name"),
                domain=cookie.get("domain"),
                path=cookie.get("path"),
            )
            removed += 1
        except TypeError:
            log("当前 Playwright 版本不支持按名称清理 Cookie，跳过自动 Cookie 清理。")
            return
        except Exception as exc:
            log(f"清理 Cookie {cookie.get('name')} 失败：{exc}")

    if removed:
        log(f"已清理 WMS 域名下疑似埋点/超长 Cookie {removed} 个。")
    else:
        log("未发现需要清理的 WMS 域名埋点/超长 Cookie。")


def launch_visible_browser(playwright, headless: bool):
    launch_options = {"headless": headless}
    attempts = [
        ("Playwright Chromium", lambda: playwright.chromium.launch(**launch_options)),
        ("本机 Chrome", lambda: playwright.chromium.launch(channel="chrome", **launch_options)),
        ("本机 Edge", lambda: playwright.chromium.launch(channel="msedge", **launch_options)),
    ]

    last_error: Exception | None = None
    for name, launcher in attempts:
        try:
            log(f"尝试启动浏览器：{name}")
            browser = launcher()
            log(f"已启动浏览器：{name}")
            return browser
        except PlaywrightError as exc:
            last_error = exc
            log(f"{name} 启动失败，继续尝试下一个浏览器。")

    raise RuntimeError(
        "没有可用的 Playwright 浏览器。请先执行：\n"
        "  python -m playwright install chromium\n"
        "或者安装 Google Chrome / Microsoft Edge 后再运行脚本。\n"
        f"最后一次错误：{last_error}"
    )


def safe_goto(page: Page, url: str, label: str, attempts: int = 3) -> None:
    for index in range(1, attempts + 1):
        try:
            log(f"{label}：{url}")
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            return
        except PlaywrightError as exc:
            message = str(exc)
            same_url_interrupt = "interrupted by another navigation" in message and url in message
            if same_url_interrupt:
                log("页面正在自行跳转到同一个地址，等待页面稳定后继续。")
                page.wait_for_timeout(3000)
                if url in page.url:
                    return
            if index == attempts:
                raise
            log(f"打开页面失败，第 {index} 次重试前等待 3 秒：{message.splitlines()[0]}")
            page.wait_for_timeout(3000)


def wait_for_page_settle(page: Page, timeout: int = 15000) -> None:
    try:
        page.wait_for_load_state("networkidle", timeout=timeout)
    except PlaywrightTimeoutError:
        log("页面可能存在后台轮询，networkidle 等待超时；继续执行。")


def first_visible_locator(page: Page, selectors: list[str], timeout: int = 1500):
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            locator.wait_for(state="visible", timeout=timeout)
            return locator
        except PlaywrightTimeoutError:
            continue
    return None


def click_by_text(page: Page, texts: list[str], exact: bool = True, timeout: int = 1500) -> bool:
    for text in texts:
        try:
            locator = page.get_by_text(text, exact=exact).first
            locator.wait_for(state="visible", timeout=timeout)
            locator.click(timeout=timeout)
            return True
        except PlaywrightError:
            continue
    return False


def click_dom_text(page: Page, text: str) -> bool:
    try:
        return bool(
            page.evaluate(
                """(targetText) => {
                    const nodes = Array.from(document.querySelectorAll("button, a, span, div, td, li"));
                    const node = nodes.find((element) => {
                        const text = (element.textContent || "").trim();
                        return text === targetText || text.includes(targetText);
                    });
                    if (!node) {
                        return false;
                    }
                    const clickable = node.closest("button, a, [role=button], .ant-select-item, .ant-btn") || node;
                    clickable.dispatchEvent(new MouseEvent("click", {
                        bubbles: true,
                        cancelable: true,
                        view: window
                    }));
                    return true;
                }""",
                text,
            )
        )
    except PlaywrightError:
        return False


def get_password(args: argparse.Namespace) -> str:
    if args.password:
        return args.password
    if args.password_env:
        value = os.environ.get(args.password_env, "")
        if value:
            return value
    if args.password_prompt:
        return getpass.getpass("请输入 WMS 密码：")
    return ""


def save_storage_state(context, storage_state: str) -> None:
    if not storage_state:
        return
    try:
        context.storage_state(path=storage_state)
        log(f"已保存登录态：{Path(storage_state).resolve()}")
    except Exception as exc:
        log(f"保存登录态失败：{exc}")


def auto_login(page: Page, username: str, password: str) -> bool:
    if not username or not password:
        log("未提供账号或密码，跳过自动登录。")
        return False

    username_input = first_visible_locator(
        page,
        [
            'input[name="username"]',
            'input[name="loginAccount"]',
            'input[name="account"]',
            'input[placeholder*="账号"]',
            'input[placeholder*="用户名"]',
            'input[placeholder*="用户"]',
            'input[type="text"]',
        ],
        timeout=3000,
    )
    password_input = first_visible_locator(
        page,
        [
            'input[name="password"]',
            'input[placeholder*="密码"]',
            'input[type="password"]',
        ],
        timeout=3000,
    )
    if username_input is None or password_input is None:
        log("没有找到登录输入框，可能已经登录或登录页结构变化。")
        return False

    log("检测到登录页，开始自动填写账号密码。")
    username_input.fill(username)
    password_input.fill(password)

    if not click_by_text(page, ["登录", "登 录", "Sign in", "Login"], exact=False, timeout=3000):
        submit = first_visible_locator(page, ['button[type="submit"]', ".ant-btn-primary"], timeout=2000)
        if submit is None:
            log("未找到登录按钮。")
            return False
        submit.click(timeout=3000)

    page.wait_for_timeout(5000)
    log("已提交登录，等待页面跳转或仓库选择。")
    return True


def auto_select_warehouse(page: Page, warehouse: str) -> bool:
    if not warehouse:
        return False

    log(f"尝试自动选择仓库：{warehouse}")
    selected = False
    if click_by_text(page, [warehouse], exact=False, timeout=3000):
        selected = True
    elif click_dom_text(page, warehouse):
        selected = True
    else:
        selector = first_visible_locator(
            page,
            [
                'input[placeholder*="仓库"]',
                'input[placeholder*="请选择"]',
                ".ant-select-selection-search-input",
            ],
            timeout=2000,
        )
        if selector is not None:
            selector.fill(warehouse)
            page.wait_for_timeout(1000)
            selected = click_by_text(page, [warehouse], exact=False, timeout=3000) or click_dom_text(page, warehouse)

    if not selected:
        log("没有自动找到仓库选项，后续如果页面未进入列表页，需要手动选择。")
        return False

    page.wait_for_timeout(1000)
    click_by_text(page, ["确定", "确认", "进入", "保存", "OK"], exact=False, timeout=2000)
    page.wait_for_timeout(3000)
    log("已尝试完成仓库选择。")
    return True


def ensure_login_and_warehouse(page: Page, context, args: argparse.Namespace) -> None:
    if args.auto_login:
        password = get_password(args)
        auto_login(page, args.username, password)
        auto_select_warehouse(page, args.warehouse)
        cleanup_tracking_cookies(context)
        save_storage_state(context, args.storage_state)
        safe_goto(page, TARGET_PAGE, "自动登录后打开订单列表页")
        wait_for_page_settle(page)
        return

    if args.no_prompt:
        cleanup_tracking_cookies(context)
        safe_goto(page, TARGET_PAGE, "无人工模式下打开订单列表页")
        wait_for_page_settle(page)
        return

    log("如果当前页面没有登录，请在打开的浏览器中手动登录。")
    log("如果登录后白屏，请手动刷新；如果出现仓库选择，请先选好仓库，直到能看到 WMS 页面。")
    input("登录、刷新白屏、选择仓库都完成后，回到终端按回车继续：")
    cleanup_tracking_cookies(context)
    save_storage_state(context, args.storage_state)


def build_list_payload(current: int, size: int, start_time: str, end_time: str, wh_code: str, status: str, channel: str = "") -> dict[str, Any]:
    return {
        "appendixFlag": "",
        "areaCodes": [],
        "categoryIdList": [],
        "cellNos": [],
        "codeType": "barcode",
        "countKind": "orderWeight",
        "countryRegionCodes": "",
        "current": current,
        "customerCodes": "",
        "endTime": end_time,
        "expressFlag": "",
        "expressPrintStatus": "",
        "forecastStatus": "",
        "logisticsCarrier": "",
        "logisticsChannel": channel,
        "orderCount": "",
        "orderNoType": "sourceNo",
        "orderSourceList": [],
        "productPackType": "",
        "receiver": "",
        "relatedReturnOrder": "",
        "salesPlatform": "",
        "size": size,
        "skuQtyStrList": [],    
        "sourceNoLists": [],
        "startTime": start_time,
        "status": status,
        "timeType": "createTime",
        "unitMark": 0,
        "varietyType": "",
        "weightCountEnd": "",
        "weightCountStart": "",
        "whCode": wh_code,
        "withVas": "",
    }


def yesterday_range() -> tuple[str, str]:
    day = datetime.now() - timedelta(days=1)
    date_text = day.strftime("%Y-%m-%d")
    return f"{date_text} 00:00:00", f"{date_text} 23:59:59"


def date_range_for_day(date_text: str) -> tuple[str, str]:
    return f"{date_text} 00:00:00", f"{date_text} 23:59:59"


def resolve_time_range(args: argparse.Namespace) -> tuple[str, str]:
    if args.start_time and args.end_time:
        return args.start_time, args.end_time
    if args.date:
        return date_range_for_day(args.date)
    return yesterday_range()


def stored_wh_code_from_file(storage_state: str) -> str:
    if not storage_state:
        return ""
    path = Path(storage_state)
    if not path.exists():
        return ""
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return ""
    local_storage = local_storage_from_state(state)
    wh_info = parse_json_text(local_storage.get("wh", ""))
    return str(wh_info.get("whCode") or "")


def resolve_wh_codes(args: argparse.Namespace) -> list[str]:
    codes: list[str] = []
    if args.wh_codes_file:
        path = Path(args.wh_codes_file)
        if not path.exists():
            raise RuntimeError(f"仓库编码文件不存在：{path}")
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                codes.append(line)

    if args.wh_codes:
        codes.extend(code.strip() for code in args.wh_codes.split(",") if code.strip())

    if not codes and args.wh_code:
        codes.append(args.wh_code)

    if not codes:
        stored_wh_code = stored_wh_code_from_file(args.storage_state)
        if stored_wh_code:
            codes.append(stored_wh_code)

    if not codes:
        codes.append("Monrovia")

    unique_codes: list[str] = []
    seen: set[str] = set()
    for code in codes:
        if code not in seen:
            unique_codes.append(code)
            seen.add(code)
    return unique_codes


def absolute_api_url(url: str) -> str:
    if url.startswith("http://") or url.startswith("https://"):
        return url
    if not url.startswith("/"):
        url = f"/{url}"
    return f"{API_ORIGIN}{url}"


def load_storage_state(path: str) -> dict[str, Any]:
    if not path:
        raise RuntimeError("没有设置 --storage-state，无法使用 HTTP 轻量模式。")

    storage_path = Path(path)
    if not storage_path.exists():
        raise RuntimeError(f"登录态文件不存在：{storage_path}")

    try:
        state = json.loads(storage_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"登录态文件不是有效 JSON：{storage_path}") from exc

    if not isinstance(state, dict):
        raise RuntimeError(f"登录态文件结构不正确：{storage_path}")
    return state


def local_storage_from_state(state: dict[str, Any], origin: str = API_ORIGIN) -> dict[str, str]:
    result: dict[str, str] = {}
    origins = state.get("origins")
    if not isinstance(origins, list):
        return result

    for origin_entry in origins:
        if not isinstance(origin_entry, dict) or origin_entry.get("origin") != origin:
            continue
        entries = origin_entry.get("localStorage")
        if not isinstance(entries, list):
            continue
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            name = entry.get("name")
            value = entry.get("value")
            if isinstance(name, str) and isinstance(value, str):
                result[name] = value
    return result


def parse_json_text(value: str) -> dict[str, Any]:
    try:
        data = json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def auth_header_candidates(token: str, auth_scheme: str) -> list[str]:
    token = token.strip()
    if not token:
        return []
    if token.lower().startswith("bearer "):
        return [token]
    if auth_scheme == "raw":
        return [token]
    if auth_scheme == "bearer":
        return [f"Bearer {token}"]
    return [f"Bearer {token}", token]


def compact_json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _add32(left: int, right: int) -> int:
    return (left + right) & 0xFFFFFFFF


def _rol32(value: int, shift: int) -> int:
    value &= 0xFFFFFFFF
    return ((value << shift) | (value >> (32 - shift))) & 0xFFFFFFFF


def _md5_cmn(q_value: int, a_value: int, b_value: int, x_value: int, shift: int, add_value: int) -> int:
    return _add32(_rol32(_add32(_add32(a_value, q_value), _add32(x_value, add_value)), shift), b_value)


def _md5_ff(a_value: int, b_value: int, c_value: int, d_value: int, x_value: int, shift: int, add_value: int) -> int:
    return _md5_cmn((b_value & c_value) | ((~b_value) & d_value), a_value, b_value, x_value, shift, add_value)


def _md5_gg(a_value: int, b_value: int, c_value: int, d_value: int, x_value: int, shift: int, add_value: int) -> int:
    return _md5_cmn((b_value & d_value) | (c_value & (~d_value)), a_value, b_value, x_value, shift, add_value)


def _md5_hh(a_value: int, b_value: int, c_value: int, d_value: int, x_value: int, shift: int, add_value: int) -> int:
    return _md5_cmn(b_value ^ c_value ^ d_value, a_value, b_value, x_value, shift, add_value)


def _md5_ii(a_value: int, b_value: int, c_value: int, d_value: int, x_value: int, shift: int, add_value: int) -> int:
    return _md5_cmn(c_value ^ (b_value | (~d_value)), a_value, b_value, x_value, shift, add_value)


def _wms_md5_words(text: str) -> list[int]:
    data = text.encode("utf-8")
    bit_len = len(data) * 8
    words_len = 16 * (((bit_len + 64) >> 9) + 1)
    words = [0] * words_len
    for index, byte in enumerate(data):
        words[index >> 2] |= byte << ((index % 4) * 8)
    words[bit_len >> 5] |= 0x80 << (bit_len % 32)
    words[14 + (((bit_len + 64) >> 9) << 4)] = bit_len

    a_value = 0x67452301
    b_value = 0xEFCDAB89
    c_value = 0x98BADEFE
    d_value = 0x10325476

    for offset in range(0, len(words), 16):
        old_a, old_b, old_c, old_d = a_value, b_value, c_value, d_value
        block = words[offset : offset + 16]

        a_value = _md5_ff(a_value, b_value, c_value, d_value, block[0], 7, -680876936)
        d_value = _md5_ff(d_value, a_value, b_value, c_value, block[1], 12, -389564586)
        c_value = _md5_ff(c_value, d_value, a_value, b_value, block[2], 17, 606105819)
        b_value = _md5_ff(b_value, c_value, d_value, a_value, block[3], 22, -1044525330)
        a_value = _md5_ff(a_value, b_value, c_value, d_value, block[4], 7, -176418897)
        d_value = _md5_ff(d_value, a_value, b_value, c_value, block[5], 12, 1200080426)
        c_value = _md5_ff(c_value, d_value, a_value, b_value, block[6], 17, -1473231341)
        b_value = _md5_ff(b_value, c_value, d_value, a_value, block[7], 22, -45705983)
        a_value = _md5_ff(a_value, b_value, c_value, d_value, block[8], 7, 1770035416)
        d_value = _md5_ff(d_value, a_value, b_value, c_value, block[9], 12, -1958414417)
        c_value = _md5_ff(c_value, d_value, a_value, b_value, block[10], 17, -42063)
        b_value = _md5_ff(b_value, c_value, d_value, a_value, block[11], 22, -1990404162)
        a_value = _md5_ff(a_value, b_value, c_value, d_value, block[12], 7, 1804603682)
        d_value = _md5_ff(d_value, a_value, b_value, c_value, block[13], 12, -40341101)
        c_value = _md5_ff(c_value, d_value, a_value, b_value, block[14], 17, -1502002290)
        b_value = _md5_ff(b_value, c_value, d_value, a_value, block[15], 22, 1236535329)

        a_value = _md5_gg(a_value, b_value, c_value, d_value, block[1], 5, -165796510)
        d_value = _md5_gg(d_value, a_value, b_value, c_value, block[6], 9, -1069501632)
        c_value = _md5_gg(c_value, d_value, a_value, b_value, block[11], 14, 643717713)
        b_value = _md5_gg(b_value, c_value, d_value, a_value, block[0], 20, -373897302)
        a_value = _md5_gg(a_value, b_value, c_value, d_value, block[5], 5, -701558691)
        d_value = _md5_gg(d_value, a_value, b_value, c_value, block[10], 9, 38016083)
        c_value = _md5_gg(c_value, d_value, a_value, b_value, block[15], 14, -660478335)
        b_value = _md5_gg(b_value, c_value, d_value, a_value, block[4], 20, -405537848)
        a_value = _md5_gg(a_value, b_value, c_value, d_value, block[9], 5, 568446438)
        d_value = _md5_gg(d_value, a_value, b_value, c_value, block[14], 9, -1019803690)
        c_value = _md5_gg(c_value, d_value, a_value, b_value, block[3], 14, -187363961)
        b_value = _md5_gg(b_value, c_value, d_value, a_value, block[8], 20, 1163531501)
        a_value = _md5_gg(a_value, b_value, c_value, d_value, block[13], 5, -1444681467)
        d_value = _md5_gg(d_value, a_value, b_value, c_value, block[2], 9, -51403784)
        c_value = _md5_gg(c_value, d_value, a_value, b_value, block[7], 14, 1735328473)
        b_value = _md5_gg(b_value, c_value, d_value, a_value, block[12], 20, -1926607734)

        a_value = _md5_hh(a_value, b_value, c_value, d_value, block[5], 4, -378558)
        d_value = _md5_hh(d_value, a_value, b_value, c_value, block[8], 11, -2022574463)
        c_value = _md5_hh(c_value, d_value, a_value, b_value, block[11], 16, 1839030562)
        b_value = _md5_hh(b_value, c_value, d_value, a_value, block[14], 23, -35309556)
        a_value = _md5_hh(a_value, b_value, c_value, d_value, block[1], 4, -1530992060)
        d_value = _md5_hh(d_value, a_value, b_value, c_value, block[4], 11, 1272893353)
        c_value = _md5_hh(c_value, d_value, a_value, b_value, block[7], 16, -155497632)
        b_value = _md5_hh(b_value, c_value, d_value, a_value, block[10], 23, -1094730640)
        a_value = _md5_hh(a_value, b_value, c_value, d_value, block[13], 4, 681279174)
        d_value = _md5_hh(d_value, a_value, b_value, c_value, block[0], 11, -358537222)
        c_value = _md5_hh(c_value, d_value, a_value, b_value, block[3], 16, -722521979)
        b_value = _md5_hh(b_value, c_value, d_value, a_value, block[6], 23, 76029189)
        a_value = _md5_hh(a_value, b_value, c_value, d_value, block[9], 4, -640364487)
        d_value = _md5_hh(d_value, a_value, b_value, c_value, block[12], 11, -421815835)
        c_value = _md5_hh(c_value, d_value, a_value, b_value, block[15], 16, 530742520)
        b_value = _md5_hh(b_value, c_value, d_value, a_value, block[2], 23, -995338651)

        a_value = _md5_ii(a_value, b_value, c_value, d_value, block[0], 6, -198630844)
        d_value = _md5_ii(d_value, a_value, b_value, c_value, block[7], 10, 1126891415)
        c_value = _md5_ii(c_value, d_value, a_value, b_value, block[14], 15, -1416354905)
        b_value = _md5_ii(b_value, c_value, d_value, a_value, block[5], 21, -57434055)
        a_value = _md5_ii(a_value, b_value, c_value, d_value, block[12], 6, 1700485571)
        d_value = _md5_ii(d_value, a_value, b_value, c_value, block[3], 10, -1894986606)
        c_value = _md5_ii(c_value, d_value, a_value, b_value, block[10], 15, -1051523)
        b_value = _md5_ii(b_value, c_value, d_value, a_value, block[1], 21, -2054922799)
        a_value = _md5_ii(a_value, b_value, c_value, d_value, block[8], 6, 1873313359)
        d_value = _md5_ii(d_value, a_value, b_value, c_value, block[15], 10, -30611744)
        c_value = _md5_ii(c_value, d_value, a_value, b_value, block[6], 15, -1560198380)
        b_value = _md5_ii(b_value, c_value, d_value, a_value, block[13], 21, 1309151649)
        a_value = _md5_ii(a_value, b_value, c_value, d_value, block[4], 6, -145523070)
        d_value = _md5_ii(d_value, a_value, b_value, c_value, block[11], 10, -1120210379)
        c_value = _md5_ii(c_value, d_value, a_value, b_value, block[2], 15, 718787259)
        b_value = _md5_ii(b_value, c_value, d_value, a_value, block[9], 21, -343485551)

        a_value = _add32(a_value, old_a)
        b_value = _add32(b_value, old_b)
        c_value = _add32(c_value, old_c)
        d_value = _add32(d_value, old_d)

    return [a_value, b_value, c_value, d_value]


def _wms_md5_bytes(text: str) -> bytes:
    return b"".join(word.to_bytes(4, "little") for word in _wms_md5_words(text))


def _wms_md5_hex(text: str) -> str:
    return _wms_md5_bytes(text).hex()


def _wms_md5_base64(text: str) -> str:
    return base64.b64encode(_wms_md5_bytes(text)).decode("ascii").rstrip("=")


def extract_js_function(source: str, marker: str) -> str:
    start = source.find(marker)
    if start < 0:
        raise RuntimeError("未能在 WMS 前端脚本中找到 Track-Key 签名函数。")

    depth = 0
    for index in range(start, len(source)):
        char = source[index]
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return source[start : index + 1]
    raise RuntimeError("WMS Track-Key 签名函数解析失败。")


def load_track_sign_js() -> str:
    global TRACK_SIGN_JS_CACHE
    if TRACK_SIGN_JS_CACHE:
        return TRACK_SIGN_JS_CACHE
    if requests is None:
        raise RuntimeError("当前 Python 环境未安装 requests，无法加载 WMS Track-Key 签名函数。")

    html = requests.get(TARGET_PAGE, timeout=30, headers={"user-agent": "Mozilla/5.0"}).text
    matches = re.findall(r'src="([^"]*/js/app\.[^"]+\.js)"', html)
    if not matches:
        raise RuntimeError("未能从 WMS 页面找到 app.js，无法生成 Track-Key。")

    app_js_url = absolute_api_url(matches[-1])
    app_js = requests.get(app_js_url, timeout=30, headers={"user-agent": "Mozilla/5.0"}).text
    TRACK_SIGN_JS_CACHE = extract_js_function(app_js, 'function g(e){const t=0,n="",i=8;')
    return TRACK_SIGN_JS_CACHE


def track_key_for_text(text: str) -> str:
    sign_function = load_track_sign_js()
    script = (
        f"const fs=require('fs');const input=fs.readFileSync(0,'utf8');"
        f"const sign={sign_function};process.stdout.write(sign(input));"
    )
    try:
        result = subprocess.run(
            ["node", "-e", script],
            input=text,
            text=True,
            capture_output=True,
            check=True,
            timeout=10,
        )
    except FileNotFoundError as exc:
        raise RuntimeError("HTTP 轻量模式生成 Track-Key 需要 node 命令，请安装 Node.js 或使用 --browser-mode。") from exc
    except subprocess.SubprocessError as exc:
        raise RuntimeError(f"生成 Track-Key 失败：{exc}") from exc
    return result.stdout.strip()


def session_from_storage_state(storage_state: str, wh_code: str, auth_scheme: str) -> tuple[Any, list[str]]:
    if requests is None:
        raise RuntimeError("当前 Python 环境未安装 requests，无法使用 HTTP 轻量模式。")

    state = load_storage_state(storage_state)
    session = requests.Session()

    for cookie in state.get("cookies", []):
        if not isinstance(cookie, dict):
            continue
        name = cookie.get("name")
        value = cookie.get("value")
        if not isinstance(name, str) or not isinstance(value, str):
            continue
        session.cookies.set(
            name,
            value,
            domain=cookie.get("domain") if isinstance(cookie.get("domain"), str) else None,
            path=cookie.get("path") if isinstance(cookie.get("path"), str) else "/",
        )

    local_storage = local_storage_from_state(state)
    wh_info = parse_json_text(local_storage.get("wh", ""))
    resolved_wh_code = wh_code or str(wh_info.get("whCode") or "")
    tenant_code = str(wh_info.get("tenantCode") or "")
    language = local_storage.get("language", "zh")
    token = local_storage.get("wms-token") or local_storage.get("omp-token") or ""
    auth_values = auth_header_candidates(token, auth_scheme)

    session.headers.update(
        {
            "accept": "application/json, text/plain, */*",
            "origin": API_ORIGIN,
            "referer": TARGET_PAGE,
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/125.0 Safari/537.36",
            "x-requested-with": "XMLHttpRequest",
        }
    )
    if resolved_wh_code:
        session.headers["whcode"] = resolved_wh_code
    if tenant_code:
        session.headers["tenantcode"] = tenant_code
    if language:
        session.headers["lang"] = language
        session.headers["language"] = language
    version = session.cookies.get("version")
    if version:
        session.headers["version"] = version
    if auth_values:
        session.headers["authorization"] = auth_values[0]

    return session, auth_values


def fetch_json_http(
    session,
    url: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    extra_headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    headers = {"accept": "application/json, text/plain, */*"}
    if extra_headers:
        headers.update(extra_headers)
    if payload is not None:
        headers["content-type"] = "application/json;charset=UTF-8"
        body = compact_json_text(payload)
        if method.upper() in {"POST", "PUT", "PATCH"}:
            headers["Track-Key"] = track_key_for_text(body)
    else:
        body = None

    response = session.request(
        method,
        absolute_api_url(url),
        data=body.encode("utf-8") if body is not None else None,
        headers=headers,
        timeout=60,
    )
    if not response.ok:
        raise RuntimeError(
            f"{method} {absolute_api_url(url)} 返回 HTTP {response.status_code} {response.reason}，"
            f"响应：{response.text[:500]}"
        )

    try:
        result = response.json()
    except ValueError as exc:
        raise RuntimeError(f"{method} {absolute_api_url(url)} 响应不是 JSON：{response.text[:500]}") from exc
    if not isinstance(result, dict):
        raise RuntimeError(f"{method} {absolute_api_url(url)} 响应 JSON 不是对象：{response.text[:500]}")
    return result


def fetch_json_http_with_auth_retry(
    session,
    auth_values: list[str],
    url: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    extra_headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    errors: list[str] = []
    candidates = auth_values or [session.headers.get("authorization", "")]
    if not candidates:
        candidates = [""]

    for index, auth_value in enumerate(candidates):
        if auth_value:
            session.headers["authorization"] = auth_value
        try:
            result = fetch_json_http(session, url, method=method, payload=payload, extra_headers=extra_headers)
            if index > 0 and auth_value in auth_values:
                auth_values.remove(auth_value)
                auth_values.insert(0, auth_value)
            return result
        except RuntimeError as exc:
            text = str(exc)
            errors.append(text)
            if "HTTP 401" not in text and "HTTP 403" not in text:
                raise

    raise RuntimeError("HTTP 轻量模式认证失败，可能登录态已过期。最后错误：" + errors[-1])


def fetch_json_in_page(
    page: Page,
    url: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    extra_headers: dict[str, str] | None = None,
) -> dict[str, Any]:
    result = page.evaluate(
        """async ({ url, method, payload, extraHeaders }) => {
            const headers = {
                "accept": "application/json, text/plain, */*",
                ...extraHeaders
            };
            const options = {
                method,
                credentials: "include",
                headers
            };
            if (payload !== null) {
                headers["content-type"] = headers["content-type"] || "application/json;charset=UTF-8";
                options.body = JSON.stringify(payload);
            }
            const response = await fetch(url, options);
            const text = await response.text();
            let json = null;
            try {
                json = text ? JSON.parse(text) : null;
            } catch (error) {
                json = null;
            }
            return {
                ok: response.ok,
                status: response.status,
                statusText: response.statusText,
                url: response.url,
                text,
                json
            };
        }""",
        {
            "url": url,
            "method": method,
            "payload": payload,
            "extraHeaders": extra_headers or {},
        },
    )

    if not result["ok"]:
        raise RuntimeError(
            f"{method} {url} 返回 HTTP {result['status']} {result['statusText']}，响应：{result['text'][:500]}"
        )
    if not isinstance(result["json"], dict):
        raise RuntimeError(f"{method} {url} 响应不是 JSON：{result['text'][:500]}")
    return result["json"]


def load_success_delivery_nos() -> set[str]:
    if not LOG_FILE.exists():
        return set()

    success: set[str] = set()
    with LOG_FILE.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        for row in reader:
            if row.get("status") == "success" and row.get("deliveryNo"):
                success.add(row["deliveryNo"])
    return success


def append_log(row: dict[str, Any]) -> None:
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    file_exists = LOG_FILE.exists()
    with LOG_FILE.open("a", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=CSV_FIELDS)
        if not file_exists:
            writer.writeheader()
        writer.writerow({field: row.get(field, "") for field in CSV_FIELDS})


def append_pdf_validation_log(row: dict[str, Any]) -> None:
    PDF_VALIDATION_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    file_exists = PDF_VALIDATION_LOG_FILE.exists()
    with PDF_VALIDATION_LOG_FILE.open("a", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=PDF_VALIDATION_FIELDS)
        if not file_exists:
            writer.writeheader()
        writer.writerow({field: row.get(field, "") for field in PDF_VALIDATION_FIELDS})


def classify_page_size(width_pt: float, height_pt: float) -> str:
    width_in = min(width_pt, height_pt) / 72
    height_in = max(width_pt, height_pt) / 72
    known_formats = [
        ("4x6", 4.0, 6.0),
        ("A6", 4.13, 5.83),
        ("A5", 5.83, 8.27),
        ("Letter", 8.5, 11.0),
        ("A4", 8.27, 11.69),
    ]
    for name, expected_width, expected_height in known_formats:
        if abs(width_in - expected_width) <= 0.25 and abs(height_in - expected_height) <= 0.25:
            return name
    return "unknown"


def validate_pdf_file(path: Path, expected_formats: set[str]) -> dict[str, Any]:
    if PdfReader is None:
        return {
            "pdfValid": False,
            "pageCount": 0,
            "pageSizes": "",
            "detectedFormats": "",
            "formatOk": False,
            "error": "当前 Python 环境未安装 pypdf",
        }

    try:
        reader = PdfReader(str(path))
        page_sizes: list[str] = []
        detected_formats: list[str] = []
        for page in reader.pages:
            width_pt = float(page.mediabox.width)
            height_pt = float(page.mediabox.height)
            width_in = width_pt / 72
            height_in = height_pt / 72
            page_sizes.append(f"{width_in:.2f}x{height_in:.2f}in")
            detected_formats.append(classify_page_size(width_pt, height_pt))

        unique_formats = sorted(set(detected_formats))
        page_count = len(reader.pages)
        format_ok = bool(page_count) and all(fmt in expected_formats for fmt in detected_formats)
        return {
            "pdfValid": bool(page_count),
            "pageCount": page_count,
            "pageSizes": "|".join(page_sizes),
            "detectedFormats": "|".join(unique_formats),
            "formatOk": format_ok,
            "error": "" if page_count else "PDF 没有页面",
        }
    except Exception as exc:
        return {
            "pdfValid": False,
            "pageCount": 0,
            "pageSizes": "",
            "detectedFormats": "",
            "formatOk": False,
            "error": str(exc),
        }


def validate_downloaded_pdfs(order: dict[str, str], paths: list[Path], expected_formats: set[str]) -> None:
    for path in paths:
        result = validate_pdf_file(path, expected_formats)
        append_pdf_validation_log(
            {
                **order,
                "filePath": str(path.resolve()),
                **result,
                "checkedAt": now_text(),
            }
        )
        status = "通过" if result["pdfValid"] and result["formatOk"] else "异常"
        log(
            f"PDF 格式检查{status}：{path.name}, "
            f"pages={result['pageCount']}, formats={result['detectedFormats'] or '-'}"
        )


def detail_url(order: dict[str, str]) -> str:
    query = urlencode(
        {
            "deliveryNo": order["deliveryNo"],
            "customerCode": order["customerCode"],
            "whCode": order["whCode"],
        }
    )
    return f"{DETAIL_API}?{query}"


def download_url_api(file_key: str, file_name: str, customer_code: str, wh_code: str) -> str:
    query = urlencode(
        {
            "fileKey": file_key,
            "fileName": file_name,
            "customerCode": customer_code,
            "whCode": wh_code,
        }
    )
    return f"{DOWNLOAD_URL_API}?{query}"


def extract_label_files(detail_data: dict[str, Any]) -> list[dict[str, Any]]:
    labels: list[dict[str, Any]] = []

    file_key = first_non_empty(detail_data, "fileKey")
    file_name = first_non_empty(detail_data, "fileName")
    if file_key:
        labels.append({"fileKey": file_key, "fileName": file_name, "raw_item": {}})
        return labels

    package_list = detail_data.get("packageList")
    if not isinstance(package_list, list):
        return labels

    seen: set[tuple[str, str]] = set()
    for package in package_list:
        if not isinstance(package, dict):
            continue
        package_file_key = first_non_empty(package, "fileKey")
        package_file_name = first_non_empty(package, "fileName")
        if not package_file_key:
            continue
        key = (package_file_key, package_file_name)
        if key in seen:
            continue
        seen.add(key)
        labels.append({"fileKey": package_file_key, "fileName": package_file_name, "raw_item": package})
    return labels


def nested_first_non_empty(data: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value: Any = data
        for part in key.split("."):
            if isinstance(value, dict):
                value = value.get(part, "")
            elif isinstance(value, list) and part.isdigit():
                index = int(part)
                value = value[index] if index < len(value) else ""
            else:
                value = ""
                break
        if value is not None and value != "":
            return str(value)
    return ""


def infer_metadata_carrier(source_fields: dict[str, str], tracking_no: str) -> str:
    text = " ".join(
        [
            source_fields.get("logisticsCarrier", ""),
            source_fields.get("logisticsChannel", ""),
            source_fields.get("logisticsChannelName", ""),
            source_fields.get("channelGroupCode", ""),
            source_fields.get("channelGroupName", ""),
            source_fields.get("fileName", ""),
            tracking_no,
        ]
    ).upper()
    carrier_keywords = [
        ("UPS", ("UPS", "1Z")),
        ("FedEx", ("FEDEX", "FED EX")),
        ("USPS", ("USPS", "92", "93", "94")),
        ("GOFO", ("GOFO", "GFUS", "YT")),
        ("SwiftX", ("SWIFTX", "SWX")),
        ("UniUni", ("UNIUNI", "UNUNU", "UUS")),
        ("SpeedX", ("SPEEDX", "SPX")),
        ("Yanwen", ("YANWEN", "YW")),
        ("OnTrac", ("ONTRAC", "D100")),
        ("LaserShip", ("LASERSHIP", "1LS")),
    ]
    for carrier, keywords in carrier_keywords:
        if any(keyword in text for keyword in keywords):
            return carrier
    return "UNKNOWN"


def infer_metadata_template(source_fields: dict[str, str]) -> tuple[str, str, str]:
    text = " ".join(
        [
            source_fields.get("logisticsChannel", ""),
            source_fields.get("logisticsChannelName", ""),
            source_fields.get("channelGroupCode", ""),
            source_fields.get("channelGroupName", ""),
            source_fields.get("fileName", ""),
            source_fields.get("fileKey", ""),
        ]
    ).upper()
    template_code = "UNKNOWN"
    if "CBT" in text:
        template_code = "CBT"
    elif "CBS" in text:
        template_code = "CBS"
    elif "0024" in text:
        template_code = "0024"

    sub_match = re.search(r"\b([AB])\s*[-_ ]?\s*([0O]?\d{1,2})\b", text)
    sub_code = ""
    if sub_match:
        digits = sub_match.group(2).replace("O", "0")
        sub_code = f"{sub_match.group(1)}{int(digits):02d}"
    return template_code, sub_code, "-".join(part for part in [template_code, sub_code] if part and part != "UNKNOWN") or "UNKNOWN"


def recognize_label_metadata(source_fields: dict[str, str], tracking_no: str) -> dict[str, str]:
    template_code, template_sub_code, template_key = infer_metadata_template(source_fields)
    carrier = infer_metadata_carrier(source_fields, tracking_no)
    channel = (
        source_fields.get("logisticsChannelName")
        or source_fields.get("logisticsChannel")
        or source_fields.get("channelGroupName")
        or source_fields.get("channelGroupCode")
        or "UNKNOWN"
    )
    known_parts = [value for value in [carrier, channel, template_code] if value and value != "UNKNOWN"]
    return {
        "carrier": carrier,
        "channel": channel,
        "template_code": template_code,
        "template_sub_code": template_sub_code or "UNKNOWN",
        "template_key": template_key,
        "confidence": "low" if known_parts else "unknown",
        "source": "metadata",
    }


def append_label_metadata(
    path: Path,
    order: dict[str, str],
    detail_data: dict[str, Any],
    label: dict[str, Any],
    url_json: dict[str, Any],
    down_load_url: str,
) -> None:
    raw_item = label.get("raw_item")
    if not isinstance(raw_item, dict):
        raw_item = {}

    source_fields = {
        "logisticsCarrier": first_non_empty(raw_item, "logisticsCarrier")
        or first_non_empty(detail_data, "logisticsCarrier"),
        "logisticsChannel": first_non_empty(detail_data, "logisticsChannel"),
        "logisticsChannelName": first_non_empty(detail_data, "logisticsChannelName"),
        "channelGroupCode": first_non_empty(detail_data, "channelGroupCode"),
        "channelGroupName": first_non_empty(detail_data, "channelGroupName"),
        "fileName": first_non_empty(label, "fileName") or first_non_empty(raw_item, "fileName"),
        "fileKey": first_non_empty(label, "fileKey") or first_non_empty(raw_item, "fileKey"),
        "logisticsSheetUrl": first_non_empty(raw_item, "logisticsSheetUrl"),
        "downLoadUrl": down_load_url,
        "deliveryNo": first_non_empty(detail_data, "deliveryNo") or order.get("deliveryNo", ""),
        "sourceNo": first_non_empty(detail_data, "sourceNo") or order.get("sourceNo", ""),
        "platformOrderNo": first_non_empty(detail_data, "platformOrderNo"),
        "referOrderNo": first_non_empty(detail_data, "referOrderNo"),
        "wmsPackNo": first_non_empty(raw_item, "wmsPackNo"),
        "customerCode": first_non_empty(detail_data, "customerCode") or order.get("customerCode", ""),
        "customerName": first_non_empty(detail_data, "customerName"),
        "whCode": first_non_empty(detail_data, "whCode") or order.get("whCode", ""),
        "whCodeName": first_non_empty(detail_data, "whCodeName"),
        "productName": nested_first_non_empty(raw_item, ["skuList.0.productName"]),
    }
    tracking_no = first_non_empty(raw_item, "expressNo") or first_non_empty(detail_data, "expressNo") or order.get("expressNo", "")
    metadata_recognition = recognize_label_metadata(source_fields, tracking_no)
    record = {
        "file_path": str(path.resolve()),
        "file_name": path.name,
        "tracking_no": tracking_no,
        "order_no": source_fields["deliveryNo"],
        "carrier_hint": source_fields["logisticsCarrier"],
        "channel_hint": source_fields["logisticsChannelName"] or source_fields["logisticsChannel"],
        "service_hint": "",
        "template_hint": "",
        "wh_code": source_fields["whCode"],
        "metadata_pre_recognition": metadata_recognition,
        "metadata_recognition": metadata_recognition,
        "source_fields": source_fields,
        "raw_item": raw_item,
        "raw_download_url_response": url_json,
        "saved_at": now_text(),
    }
    LABEL_METADATA_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with LABEL_METADATA_LOG_FILE.open("a", encoding="utf-8") as file:
        file.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")


def read_label_metadata_rows(path: Path = LABEL_METADATA_LOG_FILE) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(row, dict):
                rows.append(row)
    return rows


def flattened_metadata_row(row: dict[str, Any]) -> dict[str, Any]:
    source_fields = row.get("source_fields") if isinstance(row.get("source_fields"), dict) else {}
    recognition = (
        row.get("metadata_pre_recognition")
        if isinstance(row.get("metadata_pre_recognition"), dict)
        else row.get("metadata_recognition")
        if isinstance(row.get("metadata_recognition"), dict)
        else {}
    )
    if not recognition:
        recognition = recognize_label_metadata(
            {key: str(value or "") for key, value in source_fields.items()},
            str(row.get("tracking_no") or ""),
        )
    return {
        "file_path": row.get("file_path", ""),
        "file_name": row.get("file_name", ""),
        "tracking_no": row.get("tracking_no", ""),
        "order_no": row.get("order_no", ""),
        "wh_code": row.get("wh_code", ""),
        "logisticsCarrier": source_fields.get("logisticsCarrier", ""),
        "logisticsChannel": source_fields.get("logisticsChannel", ""),
        "logisticsChannelName": source_fields.get("logisticsChannelName", ""),
        "channelGroupCode": source_fields.get("channelGroupCode", ""),
        "channelGroupName": source_fields.get("channelGroupName", ""),
        "fileName": source_fields.get("fileName", ""),
        "carrier": recognition.get("carrier", "UNKNOWN") or "UNKNOWN",
        "channel": recognition.get("channel", "UNKNOWN") or "UNKNOWN",
        "template_code": recognition.get("template_code", "UNKNOWN") or "UNKNOWN",
        "template_sub_code": recognition.get("template_sub_code", "UNKNOWN") or "UNKNOWN",
        "template_key": recognition.get("template_key", "UNKNOWN") or "UNKNOWN",
        "confidence": recognition.get("confidence", "unknown") or "unknown",
        "saved_at": row.get("saved_at", ""),
    }


def autosize_worksheet(ws) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)


def write_label_metadata_summary(
    metadata_path: Path = LABEL_METADATA_LOG_FILE,
    summary_path: Path = LABEL_METADATA_SUMMARY_FILE,
) -> Path | None:
    rows = read_label_metadata_rows(metadata_path)
    if not rows:
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    flattened_rows = [flattened_metadata_row(row) for row in rows]
    headers = list(flattened_rows[0].keys())
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws_detail = wb.active
    ws_detail.title = "metadata_detail"
    ws_detail.append(headers)
    for item in flattened_rows:
        ws_detail.append([item.get(header, "") for header in headers])

    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws_detail.freeze_panes = "A2"
    ws_detail.auto_filter.ref = ws_detail.dimensions
    autosize_worksheet(ws_detail)

    unique_fields = [
        "logisticsCarrier",
        "logisticsChannel",
        "logisticsChannelName",
        "channelGroupCode",
        "channelGroupName",
        "fileName",
    ]
    ws_unique = wb.create_sheet("metadata_unique")
    ws_unique.append(["field", "value", "count"])
    for field in unique_fields:
        counts = Counter(str(item.get(field, "")) for item in flattened_rows)
        for value, count in sorted(counts.items(), key=lambda entry: (entry[0], entry[1])):
            ws_unique.append([field, value, count])
    for cell in ws_unique[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    autosize_worksheet(ws_unique)

    ws_recognition = wb.create_sheet("metadata_recognition")
    ws_recognition.append(["field", "value", "count"])
    for field in ["carrier", "channel", "template_code", "template_sub_code", "template_key", "confidence"]:
        counts = Counter(str(item.get(field, "UNKNOWN") or "UNKNOWN") for item in flattened_rows)
        for value, count in sorted(counts.items(), key=lambda entry: (entry[0], entry[1])):
            ws_recognition.append([field, value, count])
    for cell in ws_recognition[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    autosize_worksheet(ws_recognition)

    summary_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(summary_path)
    return summary_path


def finalize_label_metadata_outputs() -> None:
    try:
        summary_path = write_label_metadata_summary()
        if summary_path is None:
            log("没有结构化面单信息可汇总，跳过 metadata summary Excel。")
        else:
            log(f"结构化面单信息汇总 Excel: {summary_path.resolve()}")
    except Exception as exc:
        log(f"结构化面单信息汇总失败，继续原下载流程：{exc}")


def label_suffix(file_name: str) -> str:
    suffix = Path(str(file_name or "")).suffix.lower()
    return suffix if suffix in LABEL_SUFFIXES else ".pdf"


def make_label_path(order: dict[str, str], label_count: int, label_index: int, file_name: str = "") -> Path:
    delivery_no = sanitize_filename_part(order["deliveryNo"])
    express_no = sanitize_filename_part(order["expressNo"])
    if label_count > 1:
        filename = f"{delivery_no}_{express_no}_{label_index}{label_suffix(file_name)}"
    else:
        filename = f"{delivery_no}_{express_no}{label_suffix(file_name)}"
    return PDF_DIR / filename


def make_pdf_path(order: dict[str, str], label_count: int, label_index: int) -> Path:
    return make_label_path(order, label_count, label_index)


def download_pdf_with_requests(url: str, path: Path, session=None) -> None:
    if requests is None:
        raise RuntimeError("当前 Python 环境未安装 requests")

    client = session or requests
    with client.get(url, timeout=60, stream=True) as response:
        response.raise_for_status()
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("wb") as file:
            for chunk in response.iter_content(chunk_size=1024 * 256):
                if chunk:
                    file.write(chunk)

    if path.stat().st_size == 0:
        raise RuntimeError("requests 下载结果为空文件")


def download_pdf_with_playwright(context, url: str, path: Path) -> None:
    response = context.request.get(url, timeout=60000)
    if not response.ok:
        raise RuntimeError(f"Playwright 下载失败：HTTP {response.status} {response.status_text}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(response.body())
    if path.stat().st_size == 0:
        raise RuntimeError("Playwright 下载结果为空文件")


def download_pdf(context, url: str, path: Path) -> None:
    try:
        download_pdf_with_requests(url, path)
    except Exception as requests_error:
        log(f"requests 下载失败，改用 Playwright 下载：{requests_error}")
        download_pdf_with_playwright(context, url, path)


def download_pdf_http(session, url: str, path: Path) -> None:
    try:
        download_pdf_with_requests(url, path, session=session)
    except Exception as exc:
        raise RuntimeError(f"requests 下载 PDF 失败：{exc}") from exc


def normalize_order(record: dict[str, Any]) -> dict[str, str]:
    return {
        "deliveryNo": first_non_empty(record, "deliveryNo"),
        "sourceNo": first_non_empty(record, "sourceNo"),
        "customerCode": first_non_empty(record, "customerCode"),
        "whCode": first_non_empty(record, "whCode"),
        "expressNo": first_non_empty(record, "expressNo"),
    }


def fetch_order_page(
    page: Page,
    page_no: int,
    size: int,
    start_time: str,
    end_time: str,
    wh_code: str,
    status: str,
    headers_for_fetch: dict[str, str],
    channel: str = "",
) -> list[dict[str, Any]]:
    payload = build_list_payload(
        current=page_no,
        size=size,
        start_time=start_time,
        end_time=end_time,
        wh_code=wh_code,
        status=status,
        channel=channel,
    )
    list_json = fetch_json_in_page(page, LIST_API, method="POST", payload=payload, extra_headers=headers_for_fetch)
    records = list_json.get("data", {}).get("records") if isinstance(list_json.get("data"), dict) else None
    if not isinstance(records, list):
        raise RuntimeError(f"列表接口没有返回 data.records：{json.dumps(list_json, ensure_ascii=False)[:500]}")
    return records


def fetch_order_page_http(
    session,
    auth_values: list[str],
    page_no: int,
    size: int,
    start_time: str,
    end_time: str,
    wh_code: str,
    status: str,
    headers_for_fetch: dict[str, str],
    channel: str = "",
) -> list[dict[str, Any]]:
    payload = build_list_payload(
        current=page_no,
        size=size,
        start_time=start_time,
        end_time=end_time,
        wh_code=wh_code,
        status=status,
        channel=channel,
    )
    session.headers["whcode"] = wh_code
    list_json = fetch_json_http_with_auth_retry(
        session,
        auth_values,
        LIST_API,
        method="POST",
        payload=payload,
        extra_headers=headers_for_fetch,
    )
    records = list_json.get("data", {}).get("records") if isinstance(list_json.get("data"), dict) else None
    if not isinstance(records, list):
        raise RuntimeError(f"列表接口没有返回 data.records：{json.dumps(list_json, ensure_ascii=False)[:500]}")
    return records


def collect_captured_records(
    captured_list_records: dict[int, list[dict[str, Any]]],
    latest_captured_list: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[str] = set()

    for page_records in captured_list_records.values():
        records.extend(page_records)

    latest_records = latest_captured_list.get("records")
    if isinstance(latest_records, list):
        records.extend(record for record in latest_records if isinstance(record, dict))

    unique_records: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        key = first_non_empty(record, "deliveryNo", "sourceNo", "expressNo", "id") or json.dumps(
            record,
            ensure_ascii=False,
            sort_keys=True,
        )
        if key in seen:
            continue
        seen.add(key)
        unique_records.append(record)
    return unique_records


def selected_table_rows(page: Page) -> list[dict[str, str]]:
    rows = page.evaluate(
        """() => {
            const rowSet = new Set();
            const addRow = (node) => {
                const row = node && node.closest
                    ? node.closest("tr, [role=row], .ant-table-row")
                    : null;
                if (row && !row.closest("thead")) {
                    rowSet.add(row);
                }
            };

            document.querySelectorAll(".ant-table-row-selected").forEach(addRow);
            document.querySelectorAll(".ant-checkbox-checked").forEach(addRow);
            document.querySelectorAll("input[type='checkbox']:checked").forEach(addRow);
            document.querySelectorAll("[aria-selected='true']").forEach(addRow);

            return Array.from(rowSet).map((row) => ({
                rowKey: row.getAttribute("data-row-key") || row.dataset.rowKey || "",
                text: (row.innerText || row.textContent || "").replace(/\\s+/g, " ").trim()
            })).filter((row) => row.text || row.rowKey);
        }"""
    )
    if not isinstance(rows, list):
        return []
    return [
        {
            "rowKey": str(row.get("rowKey", "")),
            "text": str(row.get("text", "")),
        }
        for row in rows
        if isinstance(row, dict)
    ]


def scalar_record_values(record: dict[str, Any]) -> set[str]:
    values: set[str] = set()
    for value in record.values():
        if value is None or isinstance(value, (dict, list)):
            continue
        text = str(value).strip()
        if text:
            values.add(text)
    return values


def match_selected_records(
    selected_rows: list[dict[str, str]],
    candidate_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    matched: list[dict[str, Any]] = []
    matched_keys: set[str] = set()
    identifier_keys = ("deliveryNo", "sourceNo", "expressNo")

    for row in selected_rows:
        row_key = row.get("rowKey", "").strip()
        row_text = row.get("text", "")
        best_record: dict[str, Any] | None = None
        best_score = 0

        for record in candidate_records:
            exact_values = scalar_record_values(record)
            score = 0
            if row_key and row_key in exact_values:
                score += 10
            for key in identifier_keys:
                value = first_non_empty(record, key)
                if value and value in row_text:
                    score += 3
            if score > best_score:
                best_record = record
                best_score = score

        if best_record is None:
            continue

        dedupe_key = first_non_empty(best_record, "deliveryNo", "sourceNo", "expressNo")
        if not dedupe_key or dedupe_key in matched_keys:
            continue
        matched_keys.add(dedupe_key)
        matched.append(best_record)

    return matched


def download_order_records(
    page: Page,
    context,
    records: list[dict[str, Any]],
    extra_headers: dict[str, str],
    success_delivery_nos: set[str],
    expected_formats: set[str],
    validate_pdf: bool,
    total_limit: int,
    force: bool,
) -> tuple[int, int, int, int]:
    total_success = 0
    total_failed = 0
    total_skipped = 0
    total_seen = 0

    for record in records:
        if limit_reached(total_seen, total_limit):
            break
        total_seen += 1

        order = normalize_order(record)
        delivery_no = order["deliveryNo"]
        if not delivery_no:
            log("Skip one record without deliveryNo.")
            continue
        if not force and delivery_no in success_delivery_nos:
            total_skipped += 1
            log(f"Skip already successful order: {delivery_no}")
            continue

        last_error = ""
        for attempt in range(1, 4):
            try:
                log(f"Process order {delivery_no}, attempt {attempt}/3.")
                saved_paths = process_order(page, context, order, extra_headers)
                if validate_pdf:
                    validate_downloaded_pdfs(order, saved_paths, expected_formats)
                file_path_text = "|".join(str(path.resolve()) for path in saved_paths)
                append_log(
                    {
                        **order,
                        "status": "success",
                        "filePath": file_path_text,
                        "error": "",
                        "downloadedAt": now_text(),
                    }
                )
                success_delivery_nos.add(delivery_no)
                total_success += 1
                log(f"Order {delivery_no} downloaded: {file_path_text}")
                break
            except Exception as exc:
                last_error = str(exc)
                log(f"Order {delivery_no} failed on attempt {attempt}/3: {last_error}")
                if attempt < 3:
                    time.sleep(2)
        else:
            total_failed += 1
            append_log(
                {
                    **order,
                    "status": "failed",
                    "filePath": "",
                    "error": last_error,
                    "downloadedAt": now_text(),
                }
            )
            log(f"Order {delivery_no} finally failed, continue.")

    return total_success, total_failed, total_skipped, total_seen


def process_order_with_client(
    fetch_json: Callable[..., dict[str, Any]],
    download_file: Callable[[str, Path], None],
    order: dict[str, str],
    extra_headers: dict[str, str],
) -> list[Path]:
    detail_json = fetch_json(detail_url(order), method="GET", extra_headers=extra_headers)
    detail_data = detail_json.get("data")
    if not isinstance(detail_data, dict):
        raise RuntimeError(f"详情接口没有返回 data 对象：{json.dumps(detail_json, ensure_ascii=False)[:500]}")

    detail_order = {
        "deliveryNo": first_non_empty(detail_data, "deliveryNo") or order["deliveryNo"],
        "sourceNo": order["sourceNo"],
        "customerCode": first_non_empty(detail_data, "customerCode") or order["customerCode"],
        "whCode": first_non_empty(detail_data, "whCode") or order["whCode"],
        "expressNo": first_non_empty(detail_data, "expressNo") or order["expressNo"],
    }

    labels = extract_label_files(detail_data)
    if not labels:
        raise RuntimeError("详情接口未找到 data.fileKey，也未在 data.packageList 中找到 fileKey")

    saved_paths: list[Path] = []
    for index, label in enumerate(labels, start=1):
        file_key = label["fileKey"]
        file_name = label["fileName"] or f"{detail_order['expressNo']}.pdf"
        api_url = download_url_api(
            file_key=file_key,
            file_name=file_name,
            customer_code=detail_order["customerCode"],
            wh_code=detail_order["whCode"],
        )
        url_json = fetch_json(api_url, method="GET", extra_headers=extra_headers)
        down_load_url = url_json.get("data", {}).get("downLoadUrl") if isinstance(url_json.get("data"), dict) else ""
        if not down_load_url:
            raise RuntimeError(f"下载链接接口未返回 data.downLoadUrl：{json.dumps(url_json, ensure_ascii=False)[:500]}")

        pdf_path = make_label_path(detail_order, len(labels), index, file_name)
        log(f"开始下载面单：{pdf_path.name}")
        download_file(down_load_url, pdf_path)
        try:
            append_label_metadata(pdf_path, detail_order, detail_data, label, url_json, down_load_url)
        except Exception as exc:
            log(f"结构化面单信息写入失败，继续原下载流程：{exc}")
        saved_paths.append(pdf_path)

    return saved_paths


def process_order(page: Page, context, order: dict[str, str], extra_headers: dict[str, str]) -> list[Path]:
    def fetch_json(url: str, method: str = "GET", extra_headers: dict[str, str] | None = None, **kwargs):
        return fetch_json_in_page(page, url, method=method, extra_headers=extra_headers, **kwargs)

    def download_file(url: str, path: Path) -> None:
        download_pdf(context, url, path)

    return process_order_with_client(fetch_json, download_file, order, extra_headers)


def process_order_http(
    session,
    auth_values: list[str],
    order: dict[str, str],
    extra_headers: dict[str, str],
) -> list[Path]:
    def fetch_json(url: str, method: str = "GET", extra_headers: dict[str, str] | None = None, **kwargs):
        return fetch_json_http_with_auth_retry(
            session,
            auth_values,
            url,
            method=method,
            extra_headers=extra_headers,
            **kwargs,
        )

    def download_file(url: str, path: Path) -> None:
        download_pdf_http(session, url, path)

    return process_order_with_client(fetch_json, download_file, order, extra_headers)


def run_http_batch(
    args: argparse.Namespace,
    start_time: str,
    end_time: str,
    wh_codes: list[str],
    expected_formats: set[str],
    success_delivery_nos: set[str],
) -> tuple[int, int, int]:
    total_success = 0
    total_failed = 0
    total_skipped = 0
    total_seen = 0

    log("使用 HTTP 轻量模式：不启动浏览器，直接复用已保存登录态。")

    for wh_code in wh_codes:
        if limit_reached(total_seen, args.total_limit):
            break

        session, auth_values = session_from_storage_state(args.storage_state, wh_code, args.auth_scheme)
        log(f"开始处理仓库：{wh_code}")

        for page_no in page_numbers(args.current, args.max_pages):
            if limit_reached(total_seen, args.total_limit):
                log(f"已达到 total-limit={args.total_limit}，停止翻页。")
                break

            headers_for_fetch = {"whcode": wh_code}
            log(f"调用列表接口：current={page_no}, size={args.size}, whCode={wh_code}")
            records = fetch_order_page_http(
                session=session,
                auth_values=auth_values,
                page_no=page_no,
                size=args.size,
                start_time=start_time,
                end_time=end_time,
                wh_code=wh_code,
                status=args.statuses,
                headers_for_fetch=headers_for_fetch,
                channel=getattr(args, 'channel', ''),
            )
            log(f"列表接口返回 {len(records)} 条。")

            if not records:
                break

            for record in records:
                if limit_reached(total_seen, args.total_limit):
                    break
                total_seen += 1

                order = normalize_order(record)
                delivery_no = order["deliveryNo"]
                if not delivery_no:
                    log("跳过一条没有 deliveryNo 的记录。")
                    continue
                if not args.force and delivery_no in success_delivery_nos:
                    total_skipped += 1
                    log(f"跳过已成功订单：{delivery_no}")
                    continue

                last_error = ""
                for attempt in range(1, 4):
                    try:
                        log(f"处理订单 {delivery_no}，第 {attempt}/3 次尝试。")
                        detail_headers = {"whcode": order["whCode"] or wh_code}
                        session.headers["whcode"] = order["whCode"] or wh_code
                        saved_paths = process_order_http(session, auth_values, order, detail_headers)
                        if args.validate_pdf:
                            validate_downloaded_pdfs(order, saved_paths, expected_formats)
                        file_path_text = "|".join(str(path.resolve()) for path in saved_paths)
                        append_log(
                            {
                                **order,
                                "status": "success",
                                "filePath": file_path_text,
                                "error": "",
                                "downloadedAt": now_text(),
                            }
                        )
                        success_delivery_nos.add(delivery_no)
                        total_success += 1
                        log(f"订单 {delivery_no} 下载成功：{file_path_text}")
                        break
                    except Exception as exc:
                        last_error = str(exc)
                        log(f"订单 {delivery_no} 第 {attempt}/3 次失败：{last_error}")
                        if attempt < 3:
                            time.sleep(2)
                else:
                    total_failed += 1
                    append_log(
                        {
                            **order,
                            "status": "failed",
                            "filePath": "",
                            "error": last_error,
                            "downloadedAt": now_text(),
                        }
                    )
                    log(f"订单 {delivery_no} 最终失败，继续处理下一单。")

    log(f"处理完成：success={total_success}, failed={total_failed}, skipped={total_skipped}")
    log(f"PDF 保存目录：{PDF_DIR.resolve()}")
    log(f"下载日志：{LOG_FILE.resolve()}")
    if args.validate_pdf:
        log(f"PDF 格式检查日志：{PDF_VALIDATION_LOG_FILE.resolve()}")

    return total_success, total_failed, total_skipped


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="WMS 出库单面单批量下载")
    parser.add_argument("--current", type=int, default=1, help="起始页码，默认 1")
    parser.add_argument("--size", type=int, default=100, help="每页数量，默认 100")
    parser.add_argument("--max-pages", type=int, default=0, help="最大页数；0 表示不限，直到接口没有更多记录")
    parser.add_argument("--total-limit", type=int, default=0, help="最多处理订单数；0 表示不限")
    parser.add_argument("--headless", action="store_true", help="启用无头模式；默认 headless=False")
    parser.add_argument("--date", default="", help="目标日期 YYYY-MM-DD；不传则默认昨天")
    parser.add_argument("--start-time", default="", help="创建时间起始；优先级高于 --date")
    parser.add_argument("--end-time", default="", help="创建时间结束；优先级高于 --date")
    parser.add_argument("--wh-code", default="", help="列表查询仓库编码；不传时优先使用登录态里的当前仓库")
    parser.add_argument("--wh-codes", default="", help="多个仓库编码，逗号分隔，例如 Monrovia,US02")
    parser.add_argument("--wh-codes-file", default="", help="仓库编码文件，每行一个 whCode")
    parser.add_argument("--statuses", default="15", help="WMS 状态码，默认 15；待处理通常是 10")
    parser.add_argument("--auto-login", action="store_true", help="自动填写账号密码并选择仓库")
    parser.add_argument("--no-prompt", action="store_true", help="不等待人工输入；适合已有登录态的 headless 运行")
    parser.add_argument("--storage-state", default=str(AUTH_STATE_FILE), help="Playwright 登录态文件")
    parser.add_argument("--username", default=os.environ.get("WMS_USERNAME", ""), help="WMS 登录账号，默认读取 WMS_USERNAME")
    parser.add_argument("--password", default="", help="WMS 登录密码；不推荐使用，容易留在命令历史")
    parser.add_argument("--password-env", default="WMS_PASSWORD", help="WMS 密码环境变量名，默认 WMS_PASSWORD")
    parser.add_argument("--password-prompt", action="store_true", help="从终端隐藏输入密码")
    parser.add_argument("--warehouse", default="Monrovia", help="自动选择的仓库，默认 Monrovia")
    parser.add_argument("--validate-pdf", action="store_true", default=True, help="下载后检查 PDF 页面尺寸，默认开启")
    parser.add_argument("--no-validate-pdf", dest="validate_pdf", action="store_false", help="关闭 PDF 页面尺寸检查")
    parser.add_argument("--expected-formats", default="4x6", help="允许的 PDF 面单格式，逗号分隔，默认 4x6")
    parser.add_argument("--manual-select", action="store_true", help="手动在页面筛选并勾选订单后，只下载选中的订单")
    parser.add_argument("--browser-mode", action="store_true", help="强制使用原浏览器模式；适合首次登录或手动排查")
    parser.add_argument(
        "--auth-scheme",
        choices=["auto", "raw", "bearer"],
        default="auto",
        help="HTTP 轻量模式 Authorization 写法，默认 auto 会自动尝试 raw 和 Bearer",
    )
    parser.add_argument("--force", action="store_true", help="即使下载日志已有 success，也重新下载")
    parser.add_argument("--channel", default="", help="物流渠道筛选，例如 TikTok-CBT-US、Upload_Shipping_Label-Speedx")
    parser.add_argument("--pdf-dir", default="", help="PDF save directory override")
    parser.add_argument("--log-file", default="", help="download_log.csv path override")
    parser.add_argument("--pdf-validation-log-file", default="", help="pdf_validation_log.csv path override")
    args = parser.parse_args()
    if "--max-pages" not in sys.argv:
        args.max_pages = 0
    if "--total-limit" not in sys.argv:
        args.total_limit = 0
    return args


def main() -> None:
    global PDF_DIR, LOG_FILE, PDF_VALIDATION_LOG_FILE
    args = parse_args()
    storage_state = resolve_storage_state_path(args.storage_state)
    args.storage_state = str(storage_state)
    storage_state_exists = storage_state.exists()
    if storage_state_exists and not args.browser_mode and not args.manual_select:
        args.no_prompt = True
    if args.pdf_dir:
        PDF_DIR = Path(args.pdf_dir).expanduser().resolve()
    if args.log_file:
        LOG_FILE = Path(args.log_file).expanduser().resolve()
    if args.pdf_validation_log_file:
        PDF_VALIDATION_LOG_FILE = Path(args.pdf_validation_log_file).expanduser().resolve()
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    PDF_VALIDATION_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    log(f"当前工作目录: {CURRENT_WORK_DIR}")
    log(f"EXE 所在目录/项目目录: {BASE_DIR}")
    log(f"storage_state 实际路径: {storage_state}")
    log(f"storage_state 是否存在: {storage_state_exists}")
    start_time, end_time = resolve_time_range(args)
    wh_codes = resolve_wh_codes(args)
    expected_formats = {fmt.strip() for fmt in args.expected_formats.split(",") if fmt.strip()}

    if not wh_codes:
        raise RuntimeError("没有可用的仓库编码，请设置 --wh-code、--wh-codes 或 --wh-codes-file。")

    captured_gateway_headers: dict[str, str] = {}
    captured_list_headers: dict[str, str] = {}
    captured_list_records: dict[int, list[dict[str, Any]]] = {}
    latest_captured_list: dict[str, Any] = {"current": None, "records": None}

    success_delivery_nos = load_success_delivery_nos()
    log(f"已从日志读取 success 订单 {len(success_delivery_nos)} 个。")
    log(f"本次时间范围：{start_time} 到 {end_time}")
    log(f"本次仓库列表：{', '.join(wh_codes)}")

    if not args.browser_mode and not args.manual_select:
        try:
            run_http_batch(args, start_time, end_time, wh_codes, expected_formats, success_delivery_nos)
            finalize_label_metadata_outputs()
            return
        except Exception as exc:
            if args.no_prompt:
                raise
            log(f"HTTP 轻量模式失败，将回退到浏览器模式：{exc}")
            log("如果只是登录态过期，请在回退打开的浏览器中重新登录并按提示保存登录态。")

    if sync_playwright is None:
        raise RuntimeError("当前 Python 环境未安装 playwright，无法使用浏览器模式。请安装 playwright 或使用有效登录态运行 HTTP 轻量模式。")

    with sync_playwright() as p:
        browser = launch_visible_browser(p, headless=args.headless)
        storage_state = args.storage_state if args.storage_state and Path(args.storage_state).exists() else None
        if storage_state:
            log(f"加载已保存登录态：{Path(storage_state).resolve()}")
            context = browser.new_context(storage_state=storage_state)
        else:
            context = browser.new_context()
        page = context.new_page()

        def handle_request(request) -> None:
            if request.resource_type not in ("xhr", "fetch"):
                return
            if "/gateway/" not in request.url:
                return

            headers = reusable_headers(request.headers)
            if headers:
                captured_gateway_headers.clear()
                captured_gateway_headers.update(headers)
            if LIST_API in request.url and headers:
                captured_list_headers.clear()
                captured_list_headers.update(headers)
                log("已捕获页面真实订单列表请求头。")

        page.on("request", handle_request)

        def handle_response(response) -> None:
            request = response.request
            if request.resource_type not in ("xhr", "fetch"):
                return
            if LIST_API not in response.url:
                return

            try:
                response_json = response.json()
            except Exception:
                return

            data = response_json.get("data") if isinstance(response_json, dict) else None
            records = data.get("records") if isinstance(data, dict) else None
            current = data.get("current") if isinstance(data, dict) else None
            if response.status == 200 and isinstance(records, list):
                page_key = int(current) if isinstance(current, int) else args.current
                captured_list_records[page_key] = records
                latest_captured_list["current"] = page_key
                latest_captured_list["records"] = records
                log(f"已捕获页面真实订单列表响应：current={page_key}, {len(records)} 条。")

        page.on("response", handle_response)

        def get_captured_records(page_no: int, allow_latest_bulk: bool = False) -> list[dict[str, Any]] | None:
            if page_no in captured_list_records:
                return captured_list_records[page_no]

            latest_records = latest_captured_list.get("records")
            latest_current = latest_captured_list.get("current")
            if allow_latest_bulk and isinstance(latest_records, list):
                if len(latest_records) >= args.size or (
                    args.total_limit > 0 and len(latest_records) >= args.total_limit
                ):
                    log(
                        f"未找到 current={page_no} 的页面响应；改用最新捕获列表 "
                        f"current={latest_current}, {len(latest_records)} 条。"
                    )
                    return latest_records
            return None

        safe_goto(page, TARGET_PAGE, "打开 WMS 订单列表页")
        ensure_login_and_warehouse(page, context, args)

        safe_goto(page, TARGET_PAGE, "重新打开订单列表页")
        wait_for_page_settle(page)

        if args.manual_select:
            if args.headless or args.no_prompt:
                raise RuntimeError("--manual-select needs a visible browser and terminal prompt.")

            log("Manual select mode: filter/search in the browser, tick the rows to download, then press Enter here.")
            input("页面筛选并勾选订单后，回到终端按回车开始下载选中的 PDF：")
            page.wait_for_timeout(2000)

            selected_rows = selected_table_rows(page)
            candidate_records = collect_captured_records(captured_list_records, latest_captured_list)
            selected_records = match_selected_records(selected_rows, candidate_records)

            log(
                f"Manual select mode found {len(selected_rows)} selected table rows, "
                f"matched {len(selected_records)} order records."
            )
            if not selected_records:
                raise RuntimeError(
                    "No selected orders were matched. Please make sure the list is loaded, rows are checked, "
                    "and the selected rows are visible on the current page."
                )

            manual_headers = minimal_api_headers(captured_gateway_headers or captured_list_headers)
            total_success, total_failed, total_skipped, total_seen = download_order_records(
                page=page,
                context=context,
                records=selected_records,
                extra_headers=manual_headers,
                success_delivery_nos=success_delivery_nos,
                expected_formats=expected_formats,
                validate_pdf=args.validate_pdf,
                total_limit=args.total_limit,
                force=args.force,
            )

            log(f"处理完成：success={total_success}, failed={total_failed}, skipped={total_skipped}, seen={total_seen}")
            log(f"PDF 保存目录：{PDF_DIR.resolve()}")
            log(f"下载日志：{LOG_FILE.resolve()}")
            if args.validate_pdf:
                log(f"PDF 格式检查日志：{PDF_VALIDATION_LOG_FILE.resolve()}")
            finalize_label_metadata_outputs()
            input("按回车关闭浏览器并结束脚本：")
            browser.close()
            return

        total_success = 0
        total_failed = 0
        total_skipped = 0
        total_seen = 0

        for wh_code in wh_codes:
            if limit_reached(total_seen, args.total_limit):
                break
            log(f"开始处理仓库：{wh_code}")
            captured_list_records.clear()
            latest_captured_list["current"] = None
            latest_captured_list["records"] = None

            for page_no in page_numbers(args.current, args.max_pages):
                if limit_reached(total_seen, args.total_limit):
                    log(f"已达到 total-limit={args.total_limit}，停止翻页。")
                    break

                raw_headers_for_context = captured_list_headers or captured_gateway_headers
                headers_for_fetch = minimal_api_headers(raw_headers_for_context)
                log(f"调用列表接口：current={page_no}, size={args.size}, whCode={wh_code}")
                try:
                    records = fetch_order_page(
                        page=page,
                        page_no=page_no,
                        size=args.size,
                        start_time=start_time,
                        end_time=end_time,
                        wh_code=wh_code,
                        status=args.statuses,
                        headers_for_fetch=headers_for_fetch,
                        channel=getattr(args, 'channel', ''),
                    )
                except RuntimeError as exc:
                    header_wh_code = get_header(raw_headers_for_context, "whcode")
                    can_retry_with_header_wh = (
                        ("401" in str(exc) or "Unauthorized" in str(exc))
                        and header_wh_code
                        and header_wh_code != wh_code
                    )
                    if can_retry_with_header_wh:
                        log(f"列表接口 401，改用页面当前仓库 whcode={header_wh_code} 重试。")
                        try:
                            records = fetch_order_page(
                                page=page,
                                page_no=page_no,
                                size=args.size,
                                start_time=start_time,
                                end_time=end_time,
                                wh_code=header_wh_code,
                                status=args.statuses,
                                headers_for_fetch=headers_for_fetch,
                                channel=getattr(args, 'channel', ''),
                            )
                        except RuntimeError as retry_exc:
                            fallback_records = get_captured_records(page_no, allow_latest_bulk=True)
                            if fallback_records is not None:
                                log("evaluate(fetch) 仍然 401；改用页面自己成功返回的订单列表响应继续。")
                                records = fallback_records
                            else:
                                raise retry_exc
                    else:
                        fallback_records = get_captured_records(page_no, allow_latest_bulk=False)
                        if fallback_records is None:
                            fallback_records = get_captured_records(page_no, allow_latest_bulk=True)
                        if fallback_records is None:
                            if args.no_prompt:
                                raise
                            log("列表接口调用失败，且还没捕获到页面真实列表响应。")
                            log("请在浏览器里刷新列表页，确认仓库已选好，并手动点击查询/搜索让页面加载订单列表。")
                            input("页面订单列表正常显示后，回到终端按回车继续：")
                            page.wait_for_timeout(5000)
                            fallback_records = get_captured_records(page_no, allow_latest_bulk=True)
                        if fallback_records is None:
                            raise
                        log("evaluate(fetch) 失败；改用页面自己成功返回的订单列表响应继续。")
                        records = fallback_records
                log(f"列表接口返回 {len(records)} 条。")

                if not records:
                    break

                for record in records:
                    if limit_reached(total_seen, args.total_limit):
                        break
                    total_seen += 1

                    order = normalize_order(record)
                    delivery_no = order["deliveryNo"]
                    if not delivery_no:
                        log("跳过一条没有 deliveryNo 的记录。")
                        continue
                    if not args.force and delivery_no in success_delivery_nos:
                        total_skipped += 1
                        log(f"跳过已成功订单：{delivery_no}")
                        continue

                    last_error = ""
                    for attempt in range(1, 4):
                        try:
                            log(f"处理订单 {delivery_no}，第 {attempt}/3 次尝试。")
                            detail_headers = minimal_api_headers(captured_gateway_headers or captured_list_headers)
                            saved_paths = process_order(page, context, order, detail_headers)
                            if args.validate_pdf:
                                validate_downloaded_pdfs(order, saved_paths, expected_formats)
                            file_path_text = "|".join(str(path.resolve()) for path in saved_paths)
                            append_log(
                                {
                                    **order,
                                    "status": "success",
                                    "filePath": file_path_text,
                                    "error": "",
                                    "downloadedAt": now_text(),
                                }
                            )
                            success_delivery_nos.add(delivery_no)
                            total_success += 1
                            log(f"订单 {delivery_no} 下载成功：{file_path_text}")
                            break
                        except Exception as exc:
                            last_error = str(exc)
                            log(f"订单 {delivery_no} 第 {attempt}/3 次失败：{last_error}")
                            if attempt < 3:
                                time.sleep(2)
                    else:
                        total_failed += 1
                        append_log(
                            {
                                **order,
                                "status": "failed",
                                "filePath": "",
                                "error": last_error,
                                "downloadedAt": now_text(),
                            }
                        )
                        log(f"订单 {delivery_no} 最终失败，继续处理下一单。")

        log(f"处理完成：success={total_success}, failed={total_failed}, skipped={total_skipped}")
        log(f"PDF 保存目录：{PDF_DIR.resolve()}")
        log(f"下载日志：{LOG_FILE.resolve()}")
        if args.validate_pdf:
            log(f"PDF 格式检查日志：{PDF_VALIDATION_LOG_FILE.resolve()}")
        finalize_label_metadata_outputs()
        if not args.no_prompt:
            input("按回车关闭浏览器并结束脚本：")
        browser.close()


if __name__ == "__main__":
    main()
