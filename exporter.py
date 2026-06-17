from __future__ import annotations

import os
import json
import re
import unicodedata
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from barcode_verify_tracking import TRACKING_STATUS_STRONG, VerifyResult, match_template_text
from config import BASE_DIR
from utils import ensure_dir, log


METADATA_JSONL = Path(os.environ.get("PDF_DDD_METADATA_JSONL") or BASE_DIR / "output" / "download_label_metadata.jsonl")
UNKNOWN = "UNKNOWN"
NORMAL_LABEL_TEMPLATE = "普通面单"
SPECIAL_LABEL_TEMPLATES = {"0024", "CBT", "CBS"}
BRIEF_SHEET_NAME = "简略版"
RESULT_SHEET_NAME = "全部结果"
TEMPLATE_MISMATCH_SHEET_NAME = "下载不一致"
BRIEF_COLUMNS = ("追踪号", "承运商", "面单类型", "内容识别类型", "下载与内容对比备注", "物流渠道名称")

REVIEW_STATUSES = {
    "review_conflict",
    "review_unknown",
    "barcode_failed",
    "image_ocr_detected",
    "source_only_low_confidence",
    "ocr_needed",
    "download_file_error",
}

STATUS_ZH = {
    "auto_pass_triple_verified": "自动通过：条码、PDF文字层、下载数据一致",
    "auto_pass_barcode_verified": "自动通过：条码与下载数据一致",
    "auto_pass_text_verified": "自动通过：PDF文字层与下载数据一致",
    "auto_pass_verified_prefix": "可通过：USPS 前缀匹配",
    "source_only_low_confidence": "低置信度：仅下载数据/文件名可用",
    "review_conflict": "需要复核：多来源结果冲突",
    "review_unknown": "需要复核：无法识别",
    "barcode_failed": "条码识别失败",
    "image_ocr_detected": "图片识别到信息：需人工复核",
    "ocr_needed": "需要 OCR 深度识别",
    "download_file_error": "下载文件异常",
}

CONFIDENCE_ZH = {
    "high": "高",
    "medium": "中",
    "low": "低",
    "review": "需复核",
    "高": "高",
    "中": "中",
    "低": "低",
    "需复核": "需复核",
}

BUSINESS_COLUMNS: list[tuple[str, str, str]] = [
    ("订单号", "order_no", "raw"),
    ("追踪号", "tracking_no", "raw"),
    ("文件名", "file_name", "raw"),
    ("文件路径", "file_path", "raw"),
    ("承运商", "carrier", "raw"),
    ("面单类型", "label_template", "raw"),
    ("模板细分", "template_subdivision", "raw"),
    ("内容识别类型", "content_label_template", "raw"),
    ("内容识别依据", "content_template_source", "raw"),
    ("下载侧类型", "download_label_template", "raw"),
    ("下载侧依据", "download_template_source", "raw"),
    ("下载与内容是否一致", "template_compare_match", "yes_no"),
    ("下载与内容对比备注", "template_compare_note", "raw"),
    ("识别依据", "recognition_basis", "raw"),
    ("识别置信度", "recognition_confidence", "raw"),
    ("备注", "template_note", "raw"),
    ("是否人工复核", "need_review", "yes_no"),
    ("最终状态", "verify_status_zh", "raw"),
    ("内部状态", "verify_status", "raw"),
    ("文件类型", "file_type", "raw"),
    ("文件格式", "file_format", "raw"),
    ("条码追踪号", "barcode_tracking", "raw"),
    ("PDF文字层追踪号", "pdf_text_tracking", "raw"),
    ("文件名追踪号", "filename_tracking", "raw"),
    ("下载订单号", "download_order_no", "raw"),
    ("下载波次号", "download_wave_no", "raw"),
    ("下载仓库", "download_warehouse", "raw"),
    ("条码是否成功", "barcode_success", "yes_no"),
    ("条码候选", "all_barcode_candidates", "raw"),
    ("条码类型", "decoded_formats", "raw"),
    ("是否需要OCR", "ocr_needed", "yes_no"),
    ("OCR原因", "ocr_reason", "raw"),
    ("OCR优先级", "ocr_priority", "raw"),
    ("耗时秒", "processing_seconds", "raw"),
    ("metadata承运商", "meta_carrier", "raw"),
    ("metadata承运商冲突", "meta_carrier_conflict", "yes_no"),
    ("metadata承运商备注", "meta_carrier_note", "raw"),
    ("metadata渠道", "meta_channel", "raw"),
    ("metadata渠道代码", "meta_channel_code", "raw"),
    ("metadata客户代码", "meta_customer_code", "raw"),
    ("metadata数据来源", "meta_source", "raw"),
]


BUSINESS_COLUMNS.append(("物流渠道名称", "logistics_channel_name", "raw"))


@dataclass
class TemplateDecision:
    carrier: str = UNKNOWN
    label_template: str = NORMAL_LABEL_TEMPLATE
    template_subdivision: str = ""
    matched_text: str = ""
    matched_rule: str = ""
    source: str = "默认规则"
    confidence: str = "中"
    note: str = "未命中 0024/CBT/CBS，按普通面单处理"
    need_review: bool = False

    @property
    def recognized(self) -> bool:
        return self.label_template in SPECIAL_LABEL_TEMPLATES


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_for_match(value: object) -> str:
    text = unicodedata.normalize("NFKC", normalize_text(value)).upper()
    return re.sub(r"\s+", " ", text).strip()


def compact_text(value: object) -> str:
    text = normalize_for_match(value)
    return re.sub(r"[^0-9A-Z\u4e00-\u9fff]+", "", text)


def first_non_empty(*values: object) -> str:
    for value in values:
        text = normalize_text(value)
        if text and text.upper() != UNKNOWN:
            return text
    return ""


def confidence_to_zh(value: object, default: str = "低") -> str:
    text = normalize_text(value)
    return CONFIDENCE_ZH.get(text, default)


def status_to_zh(result: VerifyResult) -> str:
    return STATUS_ZH.get(result.verify_status, first_non_empty(result.verify_status_zh, result.verify_status))


def normalize_carrier(value: object) -> str:
    text = normalize_for_match(value)
    if not text or text == UNKNOWN:
        return ""
    rules = [
        ("USPS", ("USPS", "POSTAL")),
        ("FedEx", ("FEDEX", "FDX")),
        ("UPS", ("UPS",)),
        ("GOFO", ("GOFO", "GFUS")),
        ("SwiftX", ("SWIFTX", "SWX")),
        ("UniUni", ("UNIUNI", "UNI UNI", "UUS")),
        ("OnTrac", ("ONTRAC",)),
        ("LaserShip", ("LASERSHIP", "LASER SHIP", "1LS")),
        ("SpeedX", ("SPEEDX", "SPX")),
        ("Yanwen", ("YANWEN", "YW")),
    ]
    for carrier, tokens in rules:
        if any(token in text for token in tokens):
            return carrier
    return normalize_text(value)


def normalize_cbt_sub(raw: str) -> str:
    value = raw.upper().replace("O", "0")
    match = re.search(r"([AB])0?([1-5])", value)
    if not match:
        return ""
    return f"{match.group(1)}0{match.group(2)}"


def template_subdivision(template_code: str, template_sub_code: str, marker: str) -> str:
    if template_code == "0024":
        return template_sub_code if template_sub_code in {"01", "02"} else ""
    return ""


def find_template(text: object, source: str, confidence: str, carrier_hint: str = "") -> TemplateDecision | None:
    raw = normalize_text(text)
    if not raw:
        return None

    match = match_template_text(raw, source, confidence)
    if match.template_code:
        subdivision = template_subdivision(match.template_code, match.template_sub_code, match.template_marker)
        return TemplateDecision(
            carrier=carrier_hint or UNKNOWN,
            label_template=match.template_code,
            template_subdivision=subdivision,
            matched_text=match.template_marker,
            matched_rule=f"{match.template_code}模板规则",
            source=source,
            confidence=confidence,
            note=f"{source} 命中 {match.template_marker}",
            need_review=match.template_code == "0024" and subdivision == UNKNOWN,
        )
    return None

    normalized = normalize_for_match(raw)
    compact = compact_text(raw)

    cbt_match = re.search(r"CBT([AB][0O]?[1-5])", compact)
    if cbt_match:
        sub_code = normalize_cbt_sub(cbt_match.group(1))
        return TemplateDecision(
            carrier=carrier_hint or UNKNOWN,
            label_template="CBT",
            template_subdivision=sub_code or UNKNOWN,
            matched_text=f"CBT {sub_code}".strip(),
            matched_rule="CBT编号规则",
            source=source,
            confidence=confidence,
            note=f"{source} 命中 CBT {sub_code}" if sub_code else f"{source} 命中 CBT",
            need_review=not bool(sub_code),
        )

    if "TTCBT" in compact or re.search(r"\bTT[-_\s]*CBT\b", normalized):
        return TemplateDecision(
            carrier=carrier_hint or UNKNOWN,
            label_template="CBT",
            template_subdivision=UNKNOWN,
            matched_text="TT-CBT",
            matched_rule="TT-CBT规则",
            source=source,
            confidence=confidence,
            note=f"{source} 命中 TT-CBT，但未识别到 CBT 细分编号",
            need_review=True,
        )

    if "LAXCBS" in compact or re.search(r"\bCBS\b", normalized):
        marker = "LAX-CBS" if "LAXCBS" in compact else "CBS"
        return TemplateDecision(
            carrier=carrier_hint or UNKNOWN,
            label_template="CBS",
            template_subdivision=marker,
            matched_text=marker,
            matched_rule="CBS规则",
            source=source,
            confidence=confidence,
            note=f"{source} 命中 {marker}",
            need_review=False,
        )

    code_match = re.search(r"0024(0?[0-9]{1,2})?", compact)
    if code_match:
        raw_sub = code_match.group(1) or ""
        sub_code = ""
        if raw_sub:
            try:
                sub_code = f"{int(raw_sub.replace('O', '0')):02d}"
            except ValueError:
                sub_code = ""
        marker = f"0024 {sub_code}".strip()
        return TemplateDecision(
            carrier=carrier_hint or UNKNOWN,
            label_template="0024",
            template_subdivision=sub_code or UNKNOWN,
            matched_text=marker,
            matched_rule="0024规则",
            source=source,
            confidence=confidence,
            note=f"{source} 命中 {marker}",
            need_review=False,
        )

    return None


def read_metadata_rows(path: Path = METADATA_JSONL) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            try:
                item = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(item, dict):
                rows.append(item)
    return rows


def metadata_index(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for row in rows:
        for key in (row.get("file_path"), row.get("file_name")):
            text = normalize_text(key)
            if text:
                index[text.lower()] = row
        file_path = normalize_text(row.get("file_path"))
        if file_path:
            try:
                index[str(Path(file_path).resolve()).lower()] = row
            except OSError:
                pass
    return index


def metadata_for_result(result: VerifyResult, index: dict[str, dict[str, Any]]) -> dict[str, Any]:
    candidates = [normalize_text(result.file_path), normalize_text(result.file_name)]
    try:
        candidates.append(str(Path(result.file_path).resolve()))
    except OSError:
        pass
    for candidate in candidates:
        if candidate.lower() in index:
            return index[candidate.lower()]
    return {}


def dict_or_empty(value: object) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def flatten_named_values(prefix: str, value: object) -> Iterable[tuple[str, str]]:
    if isinstance(value, dict):
        for key, item in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from flatten_named_values(child_prefix, item)
    elif isinstance(value, list):
        for index, item in enumerate(value):
            yield from flatten_named_values(f"{prefix}[{index}]", item)
    else:
        text = normalize_text(value)
        if text:
            yield prefix, text


def metadata_named_values(metadata: dict[str, Any]) -> list[tuple[str, str]]:
    if not metadata:
        return []
    fields = dict_or_empty(metadata.get("source_fields"))
    raw_item = dict_or_empty(metadata.get("raw_item"))
    priority_keys = [
        "fileName",
        "fileKey",
        "carrier",
        "channelName",
        "logisticsChannelName",
        "deliveryChannelName",
        "serviceName",
        "logisticsChannel",
        "channelGroupCode",
        "channelGroupName",
        "productName",
    ]
    values: list[tuple[str, str]] = []
    for key in priority_keys:
        for container_name, container in (("metadata", metadata), ("metadata.source_fields", fields), ("metadata.raw_item", raw_item)):
            if key in container:
                values.append((f"{container_name}.{key}", normalize_text(container.get(key))))
    for key in ["carrier_hint", "channel_hint", "service_hint", "template_hint", "order_no", "tracking_no", "wh_code"]:
        values.append((f"metadata.{key}", normalize_text(metadata.get(key))))
    values.extend(flatten_named_values("metadata.source_fields", fields))
    values.extend(flatten_named_values("metadata.raw_item", raw_item))
    return [(name, value) for name, value in values if value]


def carrier_from_metadata(metadata: dict[str, Any]) -> str:
    for _name, value in metadata_named_values(metadata):
        carrier = normalize_carrier(value)
        if carrier:
            return carrier
    return ""


def template_from_metadata(metadata: dict[str, Any], carrier_hint: str) -> TemplateDecision | None:
    for name, value in metadata_named_values(metadata):
        decision = find_template(value, name, "高", carrier_hint)
        if decision:
            decision.note = f"{name} 命中 {decision.matched_text}"
            return decision
    return None


def template_from_filename(result: VerifyResult, carrier_hint: str) -> TemplateDecision | None:
    for source, text in [
        ("文件名", result.file_name),
        ("文件路径", result.file_path),
    ]:
        decision = find_template(text, source, "高", carrier_hint)
        if decision:
            decision.note = f"{source} 命中 {decision.matched_text}"
            return decision
    return None


def template_from_verify_result(result: VerifyResult, carrier_hint: str) -> TemplateDecision | None:
    template_code = first_non_empty(result.template_code)
    if template_code not in {"0024", "CBT", "CBS"}:
        return None
    raw_sub_code = first_non_empty(getattr(result, "template_sub_code", ""))
    marker = first_non_empty(result.template_marker, template_code)
    sub_code = template_subdivision(template_code, raw_sub_code, marker)
    display_marker = marker if template_code == "0024" else template_code
    source = first_non_empty(result.template_source, "PDF/OCR识别")
    if "OCR" in source.upper():
        confidence = "中"
    else:
        confidence = confidence_to_zh(result.template_confidence, default="高")
    decision = TemplateDecision(
        carrier=carrier_hint or UNKNOWN,
        label_template=template_code,
        template_subdivision=sub_code,
        matched_text=display_marker,
        matched_rule=f"{template_code}识别结果",
        source=source,
        confidence=confidence,
        note=f"{source} 命中 {display_marker}",
        need_review=(template_code == "0024" and sub_code == UNKNOWN),
    )
    return decision


def make_template_decision(result: VerifyResult, metadata: dict[str, Any]) -> TemplateDecision:
    carrier = first_non_empty(
        normalize_carrier(result.last_mile_carrier),
        normalize_carrier(result.carrier),
        carrier_from_metadata(metadata),
    ) or UNKNOWN

    verify_decision = template_from_verify_result(result, carrier)
    if verify_decision and "OCR" not in normalize_for_match(verify_decision.source):
        if verify_decision.carrier == UNKNOWN:
            verify_decision.carrier = carrier
        return verify_decision

    for detector in (
        lambda: template_from_metadata(metadata, carrier),
        lambda: template_from_filename(result, carrier),
    ):
        decision = detector()
        if decision:
            if decision.carrier == UNKNOWN:
                decision.carrier = carrier
            return decision

    if verify_decision:
        if verify_decision.carrier == UNKNOWN:
            verify_decision.carrier = carrier
        return verify_decision

    return TemplateDecision(
        carrier=carrier,
        label_template=NORMAL_LABEL_TEMPLATE,
        template_subdivision="",
        source="默认规则",
        confidence="中",
        note="未命中 0024/CBT/CBS，按普通面单处理",
        need_review=False,
    )


def normal_template_decision(carrier: str = UNKNOWN, source: str = "默认规则") -> TemplateDecision:
    return TemplateDecision(
        carrier=carrier or UNKNOWN,
        label_template=NORMAL_LABEL_TEMPLATE,
        template_subdivision="",
        source=source,
        confidence="中",
        note="未命中 0024/CBT/CBS，按普通面单处理",
        need_review=False,
    )


def content_template_decision(result: VerifyResult, carrier_hint: str) -> TemplateDecision:
    decision = template_from_verify_result(result, carrier_hint)
    if decision:
        return decision
    return normal_template_decision(carrier_hint, "PDF内容识别")


def normalize_template_code(value: object) -> str:
    text = normalize_for_match(value)
    if text in SPECIAL_LABEL_TEMPLATES:
        return text
    if text in {"", UNKNOWN}:
        return ""
    match = match_template_text(text, "下载metadata", "高")
    return match.template_code if match.template_code in SPECIAL_LABEL_TEMPLATES else ""


def metadata_recognition_template(metadata: dict[str, Any]) -> TemplateDecision | None:
    for key in ("metadata_recognition", "metadata_pre_recognition"):
        recognition = dict_or_empty(metadata.get(key))
        template_code = normalize_template_code(recognition.get("template_code"))
        if not template_code:
            continue
        sub_code = normalize_text(recognition.get("template_sub_code"))
        if sub_code.upper() == UNKNOWN:
            sub_code = ""
        return TemplateDecision(
            carrier=normalize_carrier(recognition.get("carrier")) or UNKNOWN,
            label_template=template_code,
            template_subdivision=template_subdivision(template_code, sub_code, normalize_text(recognition.get("template_key"))),
            matched_text=normalize_text(recognition.get("template_key") or template_code),
            matched_rule="下载metadata预识别",
            source=normalize_text(recognition.get("source")) or key,
            confidence=confidence_to_zh(recognition.get("confidence"), default="低"),
            note=f"{key}.template_code={template_code}",
            need_review=False,
        )
    return None


def download_template_decision(metadata: dict[str, Any], carrier_hint: str) -> TemplateDecision | None:
    if not metadata:
        return None

    decision = metadata_recognition_template(metadata)
    if decision:
        if decision.carrier == UNKNOWN:
            decision.carrier = carrier_hint or UNKNOWN
        return decision

    for name, value in metadata_named_values(metadata):
        decision = find_template(value, name, "高", carrier_hint)
        if decision:
            decision.note = f"{name} 命中 {decision.matched_text}"
            return decision

    return normal_template_decision(carrier_hint, "下载metadata")


def compare_templates(content: TemplateDecision, download: TemplateDecision | None) -> tuple[bool | None, str]:
    if download is None:
        return None, "无下载侧metadata，无法对比"
    if content.label_template == download.label_template:
        return True, f"一致：下载侧={download.label_template}，内容识别={content.label_template}"
    return False, f"不一致：下载侧={download.label_template}，内容识别={content.label_template}"


def tracking_no(result: VerifyResult, metadata: dict[str, Any]) -> str:
    return first_non_empty(
        result.barcode_tracking,
        result.pdf_text_tracking,
        result.source_tracking,
        result.filename_tracking,
        metadata.get("tracking_no"),
    )


def order_no(result: VerifyResult, metadata: dict[str, Any]) -> str:
    fields = dict_or_empty(metadata.get("source_fields"))
    return first_non_empty(
        result.download_order_no,
        metadata.get("order_no"),
        fields.get("deliveryNo"),
        fields.get("sourceNo"),
        fields.get("platformOrderNo"),
        fields.get("referOrderNo"),
    )


def logistics_channel_name(result: VerifyResult, metadata: dict[str, Any]) -> str:
    fields = dict_or_empty(metadata.get("source_fields"))
    return first_non_empty(
        fields.get("logisticsChannelName"),
        metadata.get("channel_hint"),
        getattr(result, "meta_channel", ""),
    )


def result_to_business_row(result: VerifyResult, metadata: dict[str, Any]) -> dict[str, object]:
    carrier = first_non_empty(
        normalize_carrier(result.last_mile_carrier),
        normalize_carrier(result.carrier),
        carrier_from_metadata(metadata),
    ) or UNKNOWN
    decision = content_template_decision(result, carrier)
    download_decision = download_template_decision(metadata, carrier)
    compare_match, compare_note = compare_templates(decision, download_decision)
    row = asdict(result)
    row.update(
        {
            "order_no": order_no(result, metadata),
            "tracking_no": tracking_no(result, metadata),
            "carrier": decision.carrier,
            "label_template": decision.label_template,
            "template_subdivision": decision.template_subdivision,
            "content_label_template": decision.label_template,
            "content_template_source": decision.source,
            "download_label_template": download_decision.label_template if download_decision else "",
            "download_template_source": download_decision.source if download_decision else "",
            "template_compare_match": compare_match,
            "template_compare_note": compare_note,
            "recognition_basis": decision.note,
            "recognition_confidence": decision.confidence,
            "template_note": decision.note,
            "template_matched_text": decision.matched_text,
            "template_matched_rule": decision.matched_rule,
            "template_source": decision.source,
            "verify_status_zh": status_to_zh(result),
            "need_review": bool(result.need_review or result.ocr_needed or decision.need_review or compare_match is False),
            "logistics_channel_name": logistics_channel_name(result, metadata),
        }
    )
    return row


def display_value(value: object, mode: str) -> object:
    if mode == "yes_no":
        if value is None:
            return ""
        return "是" if bool(value) else "否"
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return value


def export_results(results: list[VerifyResult], output_dir: Path, output_name: str) -> tuple[Path, Path]:
    ensure_dir(output_dir)
    excel_path = output_dir / f"{output_name}.xlsx"
    json_path = output_dir / f"{output_name}.json"
    write_excel(results, excel_path)
    write_json(results, json_path)
    log(f"Excel输出: {excel_path}")
    log(f"JSON输出: {json_path}")
    return excel_path, json_path


def brief_tracking_key(row: dict[str, object]) -> str:
    tracking = normalize_text(row.get("追踪号"))
    if tracking:
        return tracking.upper()
    return "|".join(normalize_text(row.get(column)) for column in BRIEF_COLUMNS).upper()


def yes_no_to_bool(value: object) -> bool:
    text = normalize_text(value)
    if text in {"是", "YES", "Yes", "yes", "TRUE", "True", "true", "1"}:
        return True
    if text in {"否", "NO", "No", "no", "FALSE", "False", "false", "0"}:
        return False
    return bool(value)


def brief_tracking_key(row: dict[str, object]) -> str:
    tracking = normalize_text(row.get("追踪号") or row.get("tracking_no"))
    if tracking:
        return tracking.upper()
    return "|".join(normalize_text(row.get(column)) for column in BRIEF_COLUMNS).upper()


def business_row_key(row: dict[str, object]) -> str:
    for key in ("tracking_no", "file_path", "order_no", "file_name"):
        value = normalize_text(row.get(key))
        if value:
            return f"{key}:{value}".upper()
    return "|".join(normalize_text(row.get(key)) for _label, key, _mode in BUSINESS_COLUMNS).upper()


def read_existing_business_rows(path: Path, sheet_name: str = RESULT_SHEET_NAME) -> list[dict[str, object]]:
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log(f"读取旧 Excel 全部结果失败，将重新生成: {path}; {exc}")
        return []
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        label_to_column = {label: (key, mode) for label, key, mode in BUSINESS_COLUMNS}
        headers = [normalize_text(value) for value in rows[0]]
        result: list[dict[str, object]] = []
        for values in rows[1:]:
            item: dict[str, object] = {}
            has_value = False
            for index, label in enumerate(headers):
                if label not in label_to_column:
                    continue
                key, mode = label_to_column[label]
                value = values[index] if index < len(values) else ""
                item[key] = yes_no_to_bool(value) if mode == "yes_no" else (value or "")
                has_value = has_value or bool(value)
            if has_value:
                result.append(item)
        return result
    finally:
        workbook.close()


def merge_business_rows(existing_rows: list[dict[str, object]], new_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    merged: list[dict[str, object]] = []
    index_by_key: dict[str, int] = {}
    for row in existing_rows:
        key = business_row_key(row)
        if key in index_by_key:
            continue
        index_by_key[key] = len(merged)
        merged.append(row)
    for row in new_rows:
        key = business_row_key(row)
        if key in index_by_key:
            merged[index_by_key[key]] = row
        else:
            index_by_key[key] = len(merged)
            merged.append(row)
    return merged


def barcode_consistent_text(row: dict[str, object]) -> str:
    if not bool(row.get("barcode_success")):
        return "否"
    if bool(row.get("barcode_matches_source")) or bool(row.get("barcode_matches_pdf_text")):
        return "是"
    if row.get("verify_status") in {"auto_pass_triple_verified", "auto_pass_barcode_verified"}:
        return "是"
    return "否"


def make_brief_row(row: dict[str, object]) -> dict[str, object]:
    return {
        "追踪号": row.get("tracking_no", ""),
        "承运商": row.get("carrier", ""),
        "面单类型": row.get("label_template", ""),
        "内容识别类型": row.get("content_label_template", ""),
        "下载与内容对比备注": row.get("template_compare_note", ""),
        "物流渠道名称": row.get("logistics_channel_name", ""),
    }


def read_existing_brief_rows(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        log(f"读取旧 Excel 简略版失败，将重新生成: {path}; {exc}")
        return []
    try:
        if BRIEF_SHEET_NAME not in workbook.sheetnames:
            return []
        sheet = workbook[BRIEF_SHEET_NAME]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_text(value) for value in rows[0]]
        index_by_name = {name: index for index, name in enumerate(headers) if name}
        result: list[dict[str, object]] = []
        for values in rows[1:]:
            item: dict[str, object] = {}
            has_value = False
            for column in BRIEF_COLUMNS:
                index = index_by_name.get(column)
                value = values[index] if index is not None and index < len(values) else ""
                item[column] = value or ""
                has_value = has_value or bool(value)
            if has_value:
                result.append(item)
        return result
    finally:
        workbook.close()


def merge_brief_rows(existing_rows: list[dict[str, object]], new_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    merged: list[dict[str, object]] = []
    index_by_key: dict[str, int] = {}
    for row in existing_rows:
        key = brief_tracking_key(row)
        if key in index_by_key:
            continue
        index_by_key[key] = len(merged)
        merged.append({column: row.get(column, "") for column in BRIEF_COLUMNS})
    for row in new_rows:
        key = brief_tracking_key(row)
        item = {column: row.get(column, "") for column in BRIEF_COLUMNS}
        if key in index_by_key:
            merged[index_by_key[key]] = item
        else:
            index_by_key[key] = len(merged)
            merged.append(item)
    return merged


def append_result_sheet(ws, rows: list[dict[str, object]]) -> None:
    ws.append([label for label, _key, _mode in BUSINESS_COLUMNS])
    for row in rows:
        ws.append([display_value(row.get(key, ""), mode) for _label, key, mode in BUSINESS_COLUMNS])
    format_result_sheet(ws)


def append_brief_sheet(ws, rows: list[dict[str, object]]) -> None:
    ws.append(list(BRIEF_COLUMNS))
    for row in rows:
        ws.append([row.get(column, "") for column in BRIEF_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def format_result_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    strong_fill = PatternFill("solid", fgColor="C6E0B4")
    weak_fill = PatternFill("solid", fgColor="FFF2CC")
    review_fill = PatternFill("solid", fgColor="F8CBAD")
    keys = [key for _label, key, _mode in BUSINESS_COLUMNS]
    status_col = keys.index("verify_status") + 1
    review_col = keys.index("need_review") + 1

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row_index in range(2, ws.max_row + 1):
        status = str(ws.cell(row_index, status_col).value or "")
        need_review = str(ws.cell(row_index, review_col).value or "") == "是"
        fill = review_fill if need_review else strong_fill if status in TRACKING_STATUS_STRONG else weak_fill
        for cell in ws[row_index]:
            cell.fill = fill

    autosize(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def count_by(rows: list[dict[str, object]], key: str) -> Counter[str]:
    return Counter(str(row.get(key) or UNKNOWN) for row in rows)


def append_counter_section(ws, title: str, headers: tuple[str, str], counter: Counter[str]) -> None:
    ws.append([])
    ws.append([title])
    ws[ws.max_row][0].font = Font(bold=True)
    ws.append(list(headers))
    for name, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])):
        ws.append([name or UNKNOWN, count])


def append_summary_sheet(ws, rows: list[dict[str, object]]) -> None:
    total = len(rows)
    auto_pass = sum(1 for row in rows if row.get("verify_status") in TRACKING_STATUS_STRONG)
    need_review = sum(1 for row in rows if bool(row.get("need_review")))
    comparable = sum(1 for row in rows if row.get("template_compare_match") is not None)
    compare_matched = sum(1 for row in rows if row.get("template_compare_match") is True)
    compare_mismatched = sum(1 for row in rows if row.get("template_compare_match") is False)

    ws.append(["统计项", "数值"])
    ws[1][0].font = Font(bold=True)
    ws[1][1].font = Font(bold=True)
    for label, value in [
        ("总文件数", total),
        ("自动通过数量", auto_pass),
        ("需要复核数量", need_review),
        ("条码识别成功数量", sum(1 for row in rows if bool(row.get("barcode_success")))),
        ("OCR触发数量", sum(1 for row in rows if bool(row.get("ocr_needed")))),
        ("下载与内容可对比数量", comparable),
        ("下载与内容一致数量", compare_matched),
        ("下载与内容不一致数量", compare_mismatched),
        ("0024数量", sum(1 for row in rows if row.get("label_template") == "0024")),
        ("CBT数量", sum(1 for row in rows if row.get("label_template") == "CBT")),
        ("CBS数量", sum(1 for row in rows if row.get("label_template") == "CBS")),
        ("普通面单数量", sum(1 for row in rows if row.get("label_template") == NORMAL_LABEL_TEMPLATE)),
    ]:
        ws.append([label, value])

    append_counter_section(ws, "按承运商统计", ("承运商", "数量"), count_by(rows, "carrier"))
    append_counter_section(ws, "按面单类型统计", ("面单类型", "数量"), count_by(rows, "label_template"))
    append_counter_section(ws, "按下载侧类型统计", ("下载侧类型", "数量"), count_by(rows, "download_label_template"))
    append_counter_section(ws, "按下载与内容是否一致统计", ("是否一致", "数量"), count_by(rows, "template_compare_match"))
    append_counter_section(ws, "按模板细分统计", ("模板细分", "数量"), count_by(rows, "template_subdivision"))

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        if row[0].row == 1 or str(row[0].value or "").startswith("按"):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    autosize(ws)


def business_rows(results: list[VerifyResult]) -> list[dict[str, object]]:
    metadata_rows = read_metadata_rows()
    index = metadata_index(metadata_rows)
    return [result_to_business_row(result, metadata_for_result(result, index)) for result in results]


def write_excel(results: list[VerifyResult], path: Path) -> None:
    current_rows = business_rows(results)
    existing_rows = read_existing_business_rows(path)
    rows = merge_business_rows(existing_rows, current_rows)
    existing_brief_rows = read_existing_brief_rows(path)
    current_brief_rows = [make_brief_row(row) for row in rows]
    brief_rows = merge_brief_rows(existing_brief_rows, current_brief_rows)

    wb = Workbook()

    ws_all = wb.active
    ws_all.title = RESULT_SHEET_NAME
    append_result_sheet(ws_all, rows)

    ws_review = wb.create_sheet("异常复核")
    append_result_sheet(ws_review, [row for row in rows if bool(row.get("need_review"))])

    ws_mismatch = wb.create_sheet(TEMPLATE_MISMATCH_SHEET_NAME)
    append_result_sheet(ws_mismatch, [row for row in rows if row.get("template_compare_match") is False])

    ws_summary = wb.create_sheet("统计汇总")
    append_summary_sheet(ws_summary, rows)

    ws_brief = wb.create_sheet(BRIEF_SHEET_NAME)
    append_brief_sheet(ws_brief, brief_rows)

    try:
        wb.save(path)
    except Exception as exc:
        raise RuntimeError(f"Excel 写入失败: {path}; {exc}") from exc


def write_json(results: list[VerifyResult], path: Path) -> None:
    metadata_rows = read_metadata_rows()
    index = metadata_index(metadata_rows)
    data: list[dict[str, object]] = []
    for result in results:
        metadata = metadata_for_result(result, index)
        carrier = first_non_empty(
            normalize_carrier(result.last_mile_carrier),
            normalize_carrier(result.carrier),
            carrier_from_metadata(metadata),
        ) or UNKNOWN
        decision = content_template_decision(result, carrier)
        download_decision = download_template_decision(metadata, carrier)
        compare_match, compare_note = compare_templates(decision, download_decision)
        row = asdict(result)
        row["business_template"] = asdict(decision)
        row["business_download_template"] = asdict(download_decision) if download_decision else {}
        row["business_template_compare_match"] = compare_match
        row["business_template_compare_note"] = compare_note
        row["business_need_review"] = bool(result.need_review or result.ocr_needed or decision.need_review or compare_match is False)
        row["business_order_no"] = order_no(result, metadata)
        row["business_tracking_no"] = tracking_no(result, metadata)
        data.append(row)
    try:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        raise RuntimeError(f"JSON 写入失败: {path}; {exc}") from exc
