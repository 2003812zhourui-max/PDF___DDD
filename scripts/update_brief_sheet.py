from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BRIEF_SHEET_NAME = "简略版"
MAIN_SHEET_NAME = "全部结果"
BRIEF_COLUMNS = ("追踪号", "承运商", "最终状态", "条码是否一致", "是否人工复核", "物流渠道名称")


def text(value: Any) -> str:
    return str(value or "").strip()


def is_yes(value: Any) -> bool:
    return text(value) in {"是", "YES", "Yes", "yes", "TRUE", "True", "true", "1"}


def autosize(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)


def sheet_rows(ws) -> list[dict[str, Any]]:
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    headers = [text(value) for value in raw_rows[0]]
    rows: list[dict[str, Any]] = []
    for values in raw_rows[1:]:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(text(value) for value in row.values()):
            rows.append(row)
    return rows


def barcode_consistent(row: dict[str, Any]) -> str:
    if "条码是否一致" in row and text(row.get("条码是否一致")):
        return text(row.get("条码是否一致"))

    tracking_no = text(row.get("追踪号"))
    barcode_tracking = text(row.get("条码追踪号"))
    status = text(row.get("最终状态"))
    barcode_success = is_yes(row.get("条码是否成功"))

    if barcode_tracking and tracking_no and barcode_tracking == tracking_no:
        return "是"
    if barcode_success and ("条码" in status or "三信号" in status or "一致" in status):
        return "是"
    return "否"


def brief_from_main(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "追踪号": row.get("追踪号", ""),
        "承运商": row.get("承运商", ""),
        "最终状态": row.get("最终状态", ""),
        "条码是否一致": barcode_consistent(row),
        "是否人工复核": row.get("是否人工复核", ""),
        "物流渠道名称": row.get("物流渠道名称", "") or row.get("metadata渠道", ""),
    }


def dedupe_key(row: dict[str, Any]) -> str:
    tracking = text(row.get("追踪号"))
    if tracking:
        return tracking.upper()
    return "|".join(text(row.get(column)) for column in BRIEF_COLUMNS).upper()


def merge_rows(existing_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in [*existing_rows, *new_rows]:
        key = dedupe_key(row)
        if key in seen:
            continue
        seen.add(key)
        merged.append({column: row.get(column, "") for column in BRIEF_COLUMNS})
    return merged


def write_brief_sheet(workbook, rows: list[dict[str, Any]], sheet_name: str = BRIEF_SHEET_NAME) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(sheet_name)
    ws.append(list(BRIEF_COLUMNS))
    for row in rows:
        ws.append([row.get(column, "") for column in BRIEF_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def update_workbook(path: Path, main_sheet: str, brief_sheet: str) -> None:
    workbook = load_workbook(path)
    if main_sheet not in workbook.sheetnames:
        raise RuntimeError(f"找不到主 sheet: {main_sheet}; 当前 sheets={workbook.sheetnames}")

    existing_rows = sheet_rows(workbook[brief_sheet]) if brief_sheet in workbook.sheetnames else []
    new_rows = [brief_from_main(row) for row in sheet_rows(workbook[main_sheet])]
    merged = merge_rows(existing_rows, new_rows)
    write_brief_sheet(workbook, merged, brief_sheet)
    workbook.save(path)
    workbook.close()
    print(f"已更新 {path} 的 {brief_sheet} sheet，记录数: {len(merged)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="增量生成或更新 Excel 的简略版 sheet")
    parser.add_argument("excel", nargs="+", help="要更新的 Excel 文件路径")
    parser.add_argument("--main-sheet", default=MAIN_SHEET_NAME, help="主 sheet 名，默认 全部结果")
    parser.add_argument("--brief-sheet", default=BRIEF_SHEET_NAME, help="简略版 sheet 名，默认 简略版")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    for item in args.excel:
        update_workbook(Path(item).expanduser().resolve(), args.main_sheet, args.brief_sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
